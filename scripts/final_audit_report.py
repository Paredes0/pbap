#!/usr/bin/env python3
"""
final_audit_report.py
======================
Aggregates all per-tool audit results into a global panoramic report.
Collects grade metrics, taxonomic bias, leakage analysis, and QC audits.
Produces JSON report, text summary, and Excel workbook.

Called by audit_pipeline.sh:
    python final_audit_report.py --config pipeline_config.yaml --output-dir Dataset_Bioactividad/Global_Audit
"""

import argparse
import glob
import json
import logging
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone

import numpy as np
import pandas as pd

from audit_lib.config import load_pipeline_config, get_all_categories, get_tools_for_category
from audit_lib.provenance import generate_provenance

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

SCRIPT_NAME = "final_audit_report.py"
SCRIPT_VERSION = "2.0.0"

GRADE_ORDER = ["Gold", "Silver", "Bronze", "Red"]


# ============================================================================
# DATA COLLECTION
# ============================================================================

def _load_json_safe(path):
    """Load a JSON file, return None if not found or invalid."""
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        log.warning(f"  Failed to load {path}: {e}")
        return None


def collect_tool_reports(base_dir, pipeline_cfg):
    """Collect all per-tool audit reports, grade metrics, and taxonomic bias."""
    tools = pipeline_cfg.get("tools", {})
    reports = {}

    for tool_id, tool_cfg in tools.items():
        tool_dir = os.path.join(base_dir, "Tool_Audits", tool_id)
        pred_dir = os.path.join(tool_dir, "predictions")

        report = {
            "tool_id": tool_id,
            "display_name": tool_cfg.get("display_name", tool_id),
            "category": tool_cfg.get("category", "unknown"),
            "stage": tool_cfg.get("stage", "unknown"),
            "status": "not_run",
        }

        # Audit report (from auditoria_validation.py)
        audit_data = _load_json_safe(os.path.join(tool_dir, f"audit_report_{tool_id}.json"))
        if audit_data:
            report["audit"] = audit_data
            report["status"] = "audited"

        # Leakage report (from cdhit_leakage_analysis.py)
        leakage_data = _load_json_safe(
            os.path.join(tool_dir, "leakage_report", f"leakage_{tool_id}_report.json"))
        if leakage_data:
            report["leakage"] = leakage_data

        # Grade metrics (from run_tool_prediction.py)
        grade_data = _load_json_safe(os.path.join(pred_dir, f"grade_metrics_{tool_id}.json"))
        if grade_data:
            report["grade_metrics"] = grade_data
            report["has_predictions"] = True

        # Taxonomic bias - Gold only (from taxonomic_bias_analysis.py)
        tax_gold = _load_json_safe(os.path.join(pred_dir, f"taxonomic_bias_{tool_id}_Gold.json"))
        if tax_gold:
            report["taxonomic_bias_gold"] = tax_gold

        # Taxonomic bias - all grades
        tax_all = _load_json_safe(os.path.join(pred_dir, f"taxonomic_bias_{tool_id}.json"))
        if tax_all:
            report["taxonomic_bias_all"] = tax_all

        # Standby
        standby_data = _load_json_safe(
            os.path.join(tool_dir, "training_data", "STANDBY_REPORT.json"))
        if standby_data:
            report["status"] = "standby"
            report["standby"] = standby_data

        reports[tool_id] = report

    return reports


# ============================================================================
# REPORT GENERATION
# ============================================================================

def generate_global_report(reports, pipeline_cfg, output_dir):
    """Generate the panoramic global audit report."""
    os.makedirs(output_dir, exist_ok=True)

    categories = get_all_categories(pipeline_cfg)

    # Status summary
    status_counts = Counter(r["status"] for r in reports.values())
    log.info("=" * 70)
    log.info("GLOBAL AUDIT REPORT")
    log.info("=" * 70)
    log.info(f"  Total tools:  {len(reports)}")
    log.info(f"  Audited:      {status_counts.get('audited', 0)}")
    log.info(f"  Standby:      {status_counts.get('standby', 0)}")
    log.info(f"  Not run:      {status_counts.get('not_run', 0)}")

    # Per-category summary
    log.info("\n--- Per-Category Summary ---")
    category_summary = {}
    for category in categories:
        cat_tools = get_tools_for_category(category, pipeline_cfg)
        cat_reports = {t: reports[t] for t in cat_tools if t in reports}

        cat_info = {
            "tools": list(cat_reports.keys()),
            "n_tools": len(cat_reports),
            "statuses": Counter(r["status"] for r in cat_reports.values()),
        }

        leakage_summary = {}
        for tid, r in cat_reports.items():
            if "leakage" in r:
                summary = r["leakage"].get("summary", {})
                grades = summary.get("grades", {})
                leakage_summary[tid] = grades

        cat_info["leakage"] = leakage_summary
        category_summary[category] = cat_info

        log.info(f"\n  Category: {category}")
        log.info(f"    Tools: {', '.join(cat_reports.keys())}")
        for tid, grades in leakage_summary.items():
            total = sum(grades.values())
            gold_pct = grades.get("Gold", 0) / total * 100 if total > 0 else 0
            red_pct = grades.get("Red", 0) / total * 100 if total > 0 else 0
            log.info(f"    {tid}: Gold={grades.get('Gold', 0)} ({gold_pct:.1f}%), "
                     f"Silver={grades.get('Silver', 0)}, "
                     f"Bronze={grades.get('Bronze', 0)}, "
                     f"Red={grades.get('Red', 0)} ({red_pct:.1f}%)")

    # Grade metrics summary
    log.info("\n--- Performance Metrics (per grade) ---")
    for tid, r in sorted(reports.items()):
        gm = r.get("grade_metrics")
        if not gm:
            continue
        metrics = gm.get("metrics", {})
        overall = metrics.get("overall", {})
        gold = metrics.get("Gold", {})
        red = metrics.get("Red", {})
        leakage_bias = (red.get("mcc", 0) or 0) - (gold.get("mcc", 0) or 0)
        log.info(f"  {tid}:")
        log.info(f"    Overall  -- Acc={overall.get('accuracy', 0):.3f}  "
                 f"Sens={overall.get('sensitivity', 0):.3f}  "
                 f"Spec={overall.get('specificity', 0):.3f}  "
                 f"MCC={overall.get('mcc', 0):.3f}")
        log.info(f"    Gold     -- Sens={gold.get('sensitivity', 0):.3f}  "
                 f"MCC={gold.get('mcc', 0):.3f}")
        log.info(f"    Red      -- Sens={red.get('sensitivity', 0):.3f}  "
                 f"MCC={red.get('mcc', 0):.3f}")
        log.info(f"    Leakage bias (Red MCC - Gold MCC) = {leakage_bias:+.4f}")
        if leakage_bias > 0.15:
            log.warning(f"    *** HIGH LEAKAGE BIAS ({leakage_bias:+.4f}) ***")

    # Taxonomic bias summary
    log.info("\n--- Taxonomic Bias Summary (Gold-only) ---")
    for tid, r in sorted(reports.items()):
        tb = r.get("taxonomic_bias_gold")
        if not tb:
            continue
        results = tb.get("results", {})
        per_group = results.get("per_taxonomic_group", {})
        if not per_group:
            continue
        log.info(f"  {tid}: {len(per_group)} taxonomic groups analyzed")
        sensitivities = []
        for grp_name, grp_data in per_group.items():
            sens = grp_data.get("sensitivity", 0)
            n = grp_data.get("n_positives", 0)
            sensitivities.append(sens)
            flag = " ***" if grp_data.get("fisher_p_bh", 1) < 0.05 else ""
            if n >= 10:
                log.info(f"    {grp_name:30s}: sens={sens:.3f} (n={n}){flag}")
        if sensitivities:
            log.info(f"    Range: {min(sensitivities):.3f} - {max(sensitivities):.3f}  "
                     f"Std: {np.std(sensitivities):.3f}")

    # Overall leakage statistics
    log.info("\n--- Overall Leakage Statistics ---")
    all_grades = {g: 0 for g in GRADE_ORDER}
    total_tested = 0

    for r in reports.values():
        if "leakage" in r:
            grades = r["leakage"].get("summary", {}).get("grades", {})
            for g in GRADE_ORDER:
                all_grades[g] += grades.get(g, 0)
            total_tested += sum(grades.values())

    if total_tested > 0:
        log.info(f"  Total test sequences across all tools: {total_tested}")
        for g in GRADE_ORDER:
            log.info(f"  {g:8s}: {all_grades[g]} ({all_grades[g]/total_tested*100:.1f}%)")

    # Quality warnings
    warnings = _collect_warnings(reports)
    log.info("\n--- Quality Warnings ---")
    if warnings:
        for w in warnings:
            log.warning(w)
    else:
        log.info("  No quality warnings.")

    # Build full report
    global_report = {
        "report_name": "Global Audit Report",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "pipeline_version": SCRIPT_VERSION,
        "total_tools": len(reports),
        "status_counts": dict(status_counts),
        "category_summary": {
            cat: {
                "tools": info["tools"],
                "statuses": dict(info["statuses"]),
                "leakage": {
                    tid: {k: int(v) for k, v in grades.items()}
                    for tid, grades in info["leakage"].items()
                },
            }
            for cat, info in category_summary.items()
        },
        "overall_leakage": {
            "total_tested": total_tested,
            **{g: all_grades[g] for g in GRADE_ORDER},
        },
        "warnings": warnings,
        "per_tool": _build_per_tool_summary(reports),
    }

    # Save JSON
    report_path = os.path.join(output_dir, "GLOBAL_AUDIT_REPORT.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(global_report, f, indent=2, ensure_ascii=False, default=str)
    log.info(f"\nGlobal report saved: {report_path}")

    # Save human-readable summary
    _write_text_summary(reports, status_counts, warnings, output_dir)

    # Save Excel workbook
    _write_excel_report(reports, pipeline_cfg, warnings, output_dir)

    return global_report


def _collect_warnings(reports):
    """Collect all quality warnings across tools."""
    warnings = []
    for tid, r in sorted(reports.items()):
        if r["status"] == "standby":
            warnings.append(f"  [STANDBY] {tid}: training data not found automatically")

        if "leakage" in r:
            grades = r["leakage"].get("summary", {}).get("grades", {})
            total = sum(grades.values())
            if total > 0:
                red_pct = grades.get("Red", 0) / total * 100
                if red_pct > 20:
                    warnings.append(f"  [HIGH LEAKAGE] {tid}: {red_pct:.1f}% Red sequences")
                gold_pct = grades.get("Gold", 0) / total * 100
                if gold_pct < 50:
                    warnings.append(f"  [LOW CONFIDENCE] {tid}: only {gold_pct:.1f}% Gold sequences")

        if "grade_metrics" in r:
            metrics = r["grade_metrics"].get("metrics", {})
            gold_mcc = metrics.get("Gold", {}).get("mcc", 0) or 0
            red_mcc = metrics.get("Red", {}).get("mcc", 0) or 0
            bias = red_mcc - gold_mcc
            if bias > 0.15:
                warnings.append(f"  [LEAKAGE BIAS] {tid}: Red MCC - Gold MCC = {bias:+.4f}")

        if "audit" in r:
            audit_data = r["audit"]
            evenness = audit_data.get("positives_taxonomy", {}).get("evenness", 1.0)
            if evenness < 0.6:
                warnings.append(f"  [BIAS] {tid}: low taxonomic evenness ({evenness:.3f})")

        if "taxonomic_bias_gold" in r:
            per_group = r["taxonomic_bias_gold"].get("results", {}).get("per_taxonomic_group", {})
            sig_groups = [g for g, d in per_group.items()
                         if d.get("fisher_p_bh", 1) < 0.05 and d.get("n_positives", 0) >= 10]
            if sig_groups:
                extras = f" (+{len(sig_groups)-3} more)" if len(sig_groups) > 3 else ""
                warnings.append(
                    f"  [TAX BIAS] {tid}: {len(sig_groups)} groups with significant bias: "
                    f"{', '.join(sig_groups[:3])}{extras}")

    return warnings


def _build_per_tool_summary(reports):
    """Build per-tool summary dict for JSON report."""
    result = {}
    for tid, r in reports.items():
        entry = {
            "display_name": r.get("display_name", tid),
            "category": r.get("category", "unknown"),
            "stage": r.get("stage", "unknown"),
            "status": r["status"],
            "leakage_grades": r.get("leakage", {}).get("summary", {}).get("grades", {}),
            "has_predictions": r.get("has_predictions", False),
        }

        # Add grade metrics if available
        if "grade_metrics" in r:
            metrics = r["grade_metrics"].get("metrics", {})
            entry["metrics_overall"] = metrics.get("overall", {})
            entry["metrics_gold"] = metrics.get("Gold", {})
            entry["metrics_red"] = metrics.get("Red", {})
            gold_mcc = metrics.get("Gold", {}).get("mcc", 0) or 0
            red_mcc = metrics.get("Red", {}).get("mcc", 0) or 0
            entry["leakage_bias"] = round(red_mcc - gold_mcc, 4)

        # Add taxonomic bias summary if available
        if "taxonomic_bias_gold" in r:
            per_group = r["taxonomic_bias_gold"].get("results", {}).get("per_taxonomic_group", {})
            entry["taxonomic_groups_tested"] = len(per_group)
            sensitivities = [d.get("sensitivity", 0) for d in per_group.values()
                            if d.get("n_positives", 0) >= 10]
            if sensitivities:
                entry["sensitivity_range"] = [round(min(sensitivities), 4),
                                              round(max(sensitivities), 4)]
                entry["sensitivity_std"] = round(float(np.std(sensitivities)), 4)

        result[tid] = entry
    return result


# ============================================================================
# TEXT SUMMARY
# ============================================================================

def _write_text_summary(reports, status_counts, warnings, output_dir):
    """Write GLOBAL_AUDIT_SUMMARY.txt."""
    summary_path = os.path.join(output_dir, "GLOBAL_AUDIT_SUMMARY.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("GLOBAL AUDIT REPORT\n")
        f.write(f"Generated: {datetime.now(timezone.utc).isoformat()}\n")
        f.write("=" * 70 + "\n\n")

        f.write(f"Total tools: {len(reports)}\n")
        f.write(f"Audited:     {status_counts.get('audited', 0)}\n")
        f.write(f"Standby:     {status_counts.get('standby', 0)}\n")
        f.write(f"Not run:     {status_counts.get('not_run', 0)}\n\n")

        # Leakage table
        f.write("-" * 70 + "\n")
        f.write(f"{'Tool':<25s} {'Category':<15s} {'Status':<10s} "
                f"{'Gold':>6s} {'Silver':>6s} {'Bronze':>6s} {'Red':>6s}\n")
        f.write("-" * 70 + "\n")
        for tid, r in sorted(reports.items()):
            grades = r.get("leakage", {}).get("summary", {}).get("grades", {})
            f.write(f"{tid:<25s} {r.get('category', 'N/A'):<15s} {r['status']:<10s} "
                    f"{grades.get('Gold', '-'):>6} {grades.get('Silver', '-'):>6} "
                    f"{grades.get('Bronze', '-'):>6} {grades.get('Red', '-'):>6}\n")

        # Performance table
        audited_with_metrics = {tid: r for tid, r in reports.items() if "grade_metrics" in r}
        if audited_with_metrics:
            f.write("\n" + "=" * 70 + "\n")
            f.write("PERFORMANCE METRICS\n")
            f.write("=" * 70 + "\n")
            f.write(f"{'Tool':<20s} {'Acc':>7s} {'Sens':>7s} {'Spec':>7s} {'MCC':>7s} "
                    f"{'Gold':>7s} {'Red':>7s} {'Bias':>8s}\n")
            f.write(f"{'':20s} {'':>7s} {'':>7s} {'':>7s} {'':>7s} "
                    f"{'MCC':>7s} {'MCC':>7s} {'(R-G)':>8s}\n")
            f.write("-" * 70 + "\n")
            for tid, r in sorted(audited_with_metrics.items()):
                m = r["grade_metrics"]["metrics"]
                ov = m.get("overall", {})
                gm_d = m.get("Gold", {})
                rm_d = m.get("Red", {})
                bias = (rm_d.get("mcc", 0) or 0) - (gm_d.get("mcc", 0) or 0)
                flag = " ***" if bias > 0.15 else ""
                f.write(f"{tid:<20s} "
                        f"{ov.get('accuracy', 0):>7.3f} "
                        f"{ov.get('sensitivity', 0):>7.3f} "
                        f"{ov.get('specificity', 0):>7.3f} "
                        f"{ov.get('mcc', 0):>7.3f} "
                        f"{gm_d.get('mcc', 0) or 0:>7.3f} "
                        f"{rm_d.get('mcc', 0) or 0:>7.3f} "
                        f"{bias:>+8.4f}{flag}\n")

        if warnings:
            f.write("\n" + "=" * 70 + "\n")
            f.write("WARNINGS\n")
            f.write("=" * 70 + "\n")
            for w in warnings:
                f.write(w + "\n")

    log.info(f"Summary saved: {summary_path}")


# ============================================================================
# EXCEL REPORT
# ============================================================================

def _write_excel_report(reports, pipeline_cfg, warnings, output_dir):
    """Generate Excel workbook with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("  openpyxl not installed -- skipping Excel report. "
                    "Install with: pip install openpyxl")
        return

    excel_path = os.path.join(output_dir, "GLOBAL_AUDIT_REPORT.xlsx")
    log.info(f"Generating Excel report: {excel_path}")

    # --- Sheet 1: Overview ---
    overview_rows = []
    for tid, r in sorted(reports.items()):
        grades = r.get("leakage", {}).get("summary", {}).get("grades", {})
        total_seqs = sum(grades.values()) if grades else 0

        row = {
            "Tool ID": tid,
            "Display Name": r.get("display_name", tid),
            "Category": r.get("category", ""),
            "Stage": r.get("stage", ""),
            "Status": r["status"],
            "Gold": grades.get("Gold", ""),
            "Silver": grades.get("Silver", ""),
            "Bronze": grades.get("Bronze", ""),
            "Red": grades.get("Red", ""),
            "Total Seqs": total_seqs if total_seqs > 0 else "",
            "Red %": round(grades.get("Red", 0) / total_seqs * 100, 1) if total_seqs > 0 else "",
        }
        overview_rows.append(row)

    df_overview = pd.DataFrame(overview_rows)

    # --- Sheet 2: Performance Metrics ---
    perf_rows = []
    for tid, r in sorted(reports.items()):
        gm = r.get("grade_metrics")
        if not gm:
            continue
        metrics = gm.get("metrics", {})
        overall = metrics.get("overall", {})
        row = {
            "Tool ID": tid,
            "Category": r.get("category", ""),
            "N Total": gm.get("n_total", ""),
            "N Positives": overall.get("n_positives", ""),
            "N Negatives": overall.get("n_negatives", ""),
        }
        for key in ["accuracy", "sensitivity", "specificity", "mcc"]:
            row[f"Overall {key.title()}"] = round(overall.get(key, 0), 4)
        for grade in GRADE_ORDER:
            gd = metrics.get(grade, {})
            row[f"{grade} N"] = gd.get("n_positives", "")
            row[f"{grade} Sens"] = round(gd.get("sensitivity", 0), 4) if gd else ""
            row[f"{grade} MCC"] = round(gd.get("mcc", 0), 4) if gd else ""
        gold_mcc = metrics.get("Gold", {}).get("mcc", 0) or 0
        red_mcc = metrics.get("Red", {}).get("mcc", 0) or 0
        row["Leakage Bias (Red-Gold MCC)"] = round(red_mcc - gold_mcc, 4)
        perf_rows.append(row)

    df_perf = pd.DataFrame(perf_rows) if perf_rows else pd.DataFrame()

    # --- Sheet 3: Taxonomic Bias ---
    tax_rows = []
    for tid, r in sorted(reports.items()):
        tb = r.get("taxonomic_bias_gold")
        if not tb:
            continue
        per_group = tb.get("results", {}).get("per_taxonomic_group", {})
        for grp_name, grp_data in sorted(per_group.items()):
            ci95 = grp_data.get("sensitivity_ci95", [0, 0])
            row = {
                "Tool ID": tid,
                "Category": r.get("category", ""),
                "Taxonomic Group": grp_name,
                "N Positives": grp_data.get("n_positives", 0),
                "TP": grp_data.get("TP", 0),
                "FN": grp_data.get("FN", 0),
                "Sensitivity": round(grp_data.get("sensitivity", 0), 4),
                "Sens CI95 Low": round(ci95[0], 4),
                "Sens CI95 High": round(ci95[1], 4),
                "MCC": round(grp_data.get("mcc", 0), 4),
                "Fisher OR": round(grp_data.get("fisher_or", 0), 4),
                "Fisher p (raw)": grp_data.get("fisher_p_raw", 1),
                "Fisher p (BH)": grp_data.get("fisher_p_bh", 1),
                "Significant": "YES" if grp_data.get("fisher_p_bh", 1) < 0.05 else "",
                "Low Power": "YES" if grp_data.get("low_power", False) else "",
            }
            tax_rows.append(row)

    df_tax = pd.DataFrame(tax_rows) if tax_rows else pd.DataFrame()

    # --- Sheet 4: QC Audit ---
    qc_rows = []
    for tid, r in sorted(reports.items()):
        audit = r.get("audit")
        if not audit:
            continue
        pos = audit.get("positives_basic", {})
        neg = audit.get("negatives_basic", {})
        tax = audit.get("positives_taxonomy", {})
        ks = audit.get("ks_test", {})
        aa = audit.get("aa_comparison", {})
        row = {
            "Tool ID": tid,
            "Category": r.get("category", ""),
            "Pos Total": pos.get("total_entries", ""),
            "Pos Unique": pos.get("unique_sequences", ""),
            "Pos Duplicates": pos.get("duplicates", ""),
            "Pos Len Mean": pos.get("length_mean", ""),
            "Pos Len Std": pos.get("length_std", ""),
            "Pos Organisms": pos.get("unique_organisms", ""),
            "Neg Total": neg.get("total_entries", ""),
            "Shannon (groups)": tax.get("shannon_groups", ""),
            "Evenness (J')": tax.get("evenness", ""),
            "Shannon (species)": tax.get("shannon_species", ""),
            "KS Statistic": ks.get("ks_statistic", ""),
            "KS p-value": ks.get("p_value", ""),
            "Chi2 AA": aa.get("chi2_vs_compare", ""),
        }
        qc_rows.append(row)

    df_qc = pd.DataFrame(qc_rows) if qc_rows else pd.DataFrame()

    # --- Sheet 5: Warnings ---
    df_warnings = pd.DataFrame({"Warning": warnings}) if warnings else pd.DataFrame()

    # --- Write Excel ---
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_overview.to_excel(writer, sheet_name="Overview", index=False)
        if not df_perf.empty:
            df_perf.to_excel(writer, sheet_name="Performance", index=False)
        if not df_tax.empty:
            df_tax.to_excel(writer, sheet_name="Taxonomic Bias", index=False)
        if not df_qc.empty:
            df_qc.to_excel(writer, sheet_name="QC Audit", index=False)
        if not df_warnings.empty:
            df_warnings.to_excel(writer, sheet_name="Warnings", index=False)

    # Apply formatting
    _format_excel(excel_path)
    log.info(f"Excel report saved: {excel_path}")


def _format_excel(excel_path):
    """Apply formatting to the Excel workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    wb = load_workbook(excel_path)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                cell.border = thin_border
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                min(max_len + 3, 30)

        if ws.title == "Overview":
            _format_overview_sheet(ws, red_fill, yellow_fill, green_fill)
        elif ws.title == "Performance":
            _format_performance_sheet(ws, red_fill, yellow_fill)
        elif ws.title == "Taxonomic Bias":
            _format_tax_sheet(ws, yellow_fill)

    wb.save(excel_path)


def _find_col_idx(ws, header_name):
    """Find column index by header name (1-based)."""
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    return None


def _format_overview_sheet(ws, red_fill, yellow_fill, green_fill):
    """Conditional formatting for Overview sheet."""
    status_col = _find_col_idx(ws, "Status")
    red_pct_col = _find_col_idx(ws, "Red %")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if status_col:
            cell = row[status_col - 1]
            if cell.value == "audited":
                cell.fill = green_fill
            elif cell.value == "standby":
                cell.fill = yellow_fill
            elif cell.value == "not_run":
                cell.fill = red_fill

        if red_pct_col:
            cell = row[red_pct_col - 1]
            if cell.value and isinstance(cell.value, (int, float)):
                if cell.value > 20:
                    cell.fill = red_fill
                elif cell.value > 10:
                    cell.fill = yellow_fill


def _format_performance_sheet(ws, red_fill, yellow_fill):
    """Conditional formatting for Performance sheet."""
    bias_col = _find_col_idx(ws, "Leakage Bias (Red-Gold MCC)")
    if bias_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[bias_col - 1]
            if cell.value and isinstance(cell.value, (int, float)):
                if cell.value > 0.15:
                    cell.fill = red_fill
                elif cell.value > 0.05:
                    cell.fill = yellow_fill


def _format_tax_sheet(ws, yellow_fill):
    """Conditional formatting for Taxonomic Bias sheet."""
    sig_col = _find_col_idx(ws, "Significant")
    if sig_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[sig_col - 1]
            if cell.value == "YES":
                for c in row:
                    c.fill = yellow_fill


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Generate global audit report.")
    parser.add_argument("--config", required=True, dest="pipeline_config",
                        help="Path to pipeline_config.yaml.")
    parser.add_argument("--output-dir", required=True,
                        help="Output directory for global report.")
    args = parser.parse_args()

    cfg = load_pipeline_config(args.pipeline_config)

    # Determine base_dir (parent of output_dir, which is Global_Audit)
    base_dir = os.path.normpath(os.path.join(args.output_dir, ".."))

    # Collect all tool reports
    reports = collect_tool_reports(base_dir, cfg)

    # Generate global report
    report = generate_global_report(reports, cfg, args.output_dir)

    # Provenance
    generate_provenance(
        output_dir=args.output_dir,
        script_name=SCRIPT_NAME,
        parameters={"total_tools": len(reports)},
        output_stats={
            "audited": sum(1 for r in reports.values() if r["status"] == "audited"),
            "standby": sum(1 for r in reports.values() if r["status"] == "standby"),
            "with_predictions": sum(1 for r in reports.values() if r.get("has_predictions")),
            "with_taxonomic_bias": sum(1 for r in reports.values() if r.get("taxonomic_bias_gold")),
        },
    )

    log.info("\n" + "=" * 70)
    log.info("GLOBAL REPORT COMPLETE")
    log.info("=" * 70)


if __name__ == "__main__":
    main()

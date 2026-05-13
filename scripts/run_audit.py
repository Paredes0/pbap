"""
Pipeline Fase 1 — orchestrator E2E (orchestrator_design.md compliant).

Reads a FASTA of peptides, runs N tools, and produces under <output>/:
  consolidated.csv         wide format, one row per peptide
  consolidated.json        nested {peptide: {predictions, extra_metrics, agreements}}
  per_tool/<tool>/         raw outputs + run log per tool
  tool_health_report.json  per-tool status / runtime / diagnosis
  REPORT.md                6-section Markdown summary

Schema (docs/orchestrator_design.md §1-§3):
  - Eje binario: class_norm in {positive, negative, None}, score in [0,1] | None.
  - Eje extra_metrics: per-tool dict {metric_name: {value: float, unit: str}}.
  - Coexisten; un tool puede llenar uno, otro o ambos. apex -> sólo extra_metrics (34 MIC strains).

Score sanity (binario):
  Asumimos cada tool emite probabilidad calibrada en [0,1]. NO normalizamos
  silenciosamente — transformaciones erróneas son peores que dejar el dato crudo.
  Si un tool emite valores fuera de [0,1] se cuenta y se reporta en
  tool_health_report.score_out_of_range; el valor original se preserva.

Agreement intra-categoría (Opción B, §4):
  Por cada categoría con >=2 tools binarios definidos (en este run), se añade
  columna agreement_<categoria>:
    consensus_positive | consensus_negative | split | single_tool
  Sin voting, sin promedio, sin weighted ensemble (Opción E diferida).
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
PROJECT_ROOT = REPO_ROOT
INPUTS_DIR = PROJECT_ROOT / "Inputs"
OUTPUTS_DIR = PROJECT_ROOT / "Outputs"
sys.path.insert(0, str(REPO_ROOT))

from audit_lib.config import get_tool_config, load_pipeline_config  # noqa: E402
from audit_lib.tool_runner import run_tool  # noqa: E402

DEFAULT_TOOLS = [
    "toxinpred3",
    "antibp3",
    "hemopi2",
    "hemodl",
    "deepb3p",
    "deepbp",
    "apex",
    "perseucpp",
    "acp_dpe",
    "bertaip",
]


def parse_fasta(path: Path) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    cur_header: str | None = None
    cur_seq: list[str] = []
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith(">"):
            if cur_header is not None:
                pairs.append((cur_header, "".join(cur_seq).upper()))
            cur_header = line[1:].split()[0]
            cur_seq = []
        else:
            cur_seq.append(line)
    if cur_header is not None:
        pairs.append((cur_header, "".join(cur_seq).upper()))
    return pairs


def _read_csv_rows(path: Path, header: bool) -> list[dict | list]:
    text = path.read_text(encoding="utf-8", errors="replace")
    if header:
        return list(csv.DictReader(text.splitlines()))
    return list(csv.reader(text.splitlines()))


def _normalize_id(value: str, strip_prefix: str | None) -> str:
    if value is None:
        return ""
    v = str(value).strip()
    if strip_prefix and v.startswith(strip_prefix):
        v = v[len(strip_prefix):]
    return v


def _to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _apply_transform(value: float | None, transform: str) -> float | None:
    if value is None:
        return None
    if transform in (None, "identity"):
        return value
    if transform == "log10":
        import math
        return math.log10(value) if value > 0 else None
    if transform == "reciprocal":
        return 1.0 / value if value != 0 else None
    raise ValueError(f"Unknown extra_metric transform: {transform!r}")


def _parse_predictions(
    tool_id: str,
    predictions_path: Path,
    parsing: dict,
    peptides: list[tuple[str, str]],
) -> list[dict]:
    """Return one normalized record per peptide (in input order).

    Each record = {class_norm, score, raw_class, extra_metrics, missing_in_output}.
    extra_metrics is dict[metric_name -> {value, unit}], possibly empty.
    """
    if tool_id == "deepbp":
        return _parse_deepbp_stdout(predictions_path, peptides, parsing)

    has_header = parsing.get("header", True)
    rows = _read_csv_rows(predictions_path, has_header)

    id_col = parsing.get("id_column")
    score_col = parsing.get("score_column")
    pred_col = parsing.get("prediction_column")
    positive_label = parsing.get("positive_label")
    strip_prefix = parsing.get("id_strip_prefix")
    id_is_sequence = parsing.get("id_is_sequence", False)
    score_threshold = parsing.get("score_threshold")
    pos_class = parsing.get("positive_class_label", "positive")
    neg_class = parsing.get("negative_class_label", "negative")
    prediction_type = parsing.get("prediction_type", "classification")
    prefer_threshold = bool(parsing.get("prefer_threshold_over_raw_class", False))
    extras_spec = parsing.get("extra_metrics") or []

    if has_header:
        index: dict[str, dict] = {}
        for r in rows:
            raw_id = r.get(id_col, "")
            key = _normalize_id(raw_id, strip_prefix).upper() if id_is_sequence else _normalize_id(raw_id, strip_prefix)
            index[key] = r
    else:
        id_idx = parsing.get("id_column_index", 0)
        score_idx = parsing.get("score_column_index")
        index = {}
        for r in rows:
            if not r:
                continue
            key = _normalize_id(r[id_idx], strip_prefix)
            entry = {"_row": r, "_score": r[score_idx] if score_idx is not None else None}
            index[key] = entry

    records: list[dict] = []
    for hdr, seq in peptides:
        lookup_key = seq if id_is_sequence else hdr
        row = index.get(lookup_key)
        if row is None:
            records.append({
                "class_norm": None, "score": None, "raw_class": None,
                "extra_metrics": {}, "missing_in_output": True,
            })
            continue

        if has_header:
            score_val = _to_float(row.get(score_col)) if score_col else None
            raw_class = row.get(pred_col) if pred_col else None
        else:
            score_val = _to_float(row.get("_score"))
            raw_class = None

        class_norm = _derive_class_norm(
            raw_class, score_val, positive_label, score_threshold,
            pos_class, neg_class, prediction_type,
            prefer_threshold=prefer_threshold,
        )

        extras: dict[str, dict] = {}
        if has_header:
            for spec in extras_spec:
                fld = spec.get("field")
                val = _to_float(row.get(fld)) if fld else None
                val = _apply_transform(val, spec.get("transform", "identity"))
                extras[spec["name"]] = {"value": val, "unit": spec.get("unit", "none")}

        records.append({
            "class_norm": class_norm,
            "score": score_val,
            "raw_class": raw_class,
            "extra_metrics": extras,
            "missing_in_output": False,
        })
    return records


def _derive_class_norm(raw_class, score, positive_label, threshold,
                       pos_label, neg_label, prediction_type,
                       prefer_threshold: bool = False):
    """Determine class_norm from tool output.

    Default: si raw_class está presente y hay positive_label, usar raw_class.
    Si `prefer_threshold=True` Y hay score+threshold válidos, ignora raw_class y usa
    el threshold del YAML — útil cuando el tool emite su clasificación con threshold
    interno fijo (ej. bertaip con corte 0.5) y el usuario quiere subir el listón
    desde el YAML sin parchar el script del tool.
    """
    if prediction_type in ("regression", "extra_only"):
        return None
    if prefer_threshold and score is not None and threshold is not None:
        return "positive" if score >= threshold else "negative"
    if raw_class is not None and positive_label is not None:
        return "positive" if str(raw_class).strip() == str(positive_label) else "negative"
    if score is not None and threshold is not None:
        return "positive" if score >= threshold else "negative"
    return None


_PYTHON_LIST_LINE_RE = re.compile(r"^\[(.*)\]$")


def _parse_deepbp_stdout(predictions_path: Path, peptides, parsing) -> list[dict]:
    """deepbp prints `['ACP', 'non-ACP', ...]` on its own line near the end of stdout.

    Keras progress lines like `1/1 [==============================] - 1s 528ms/step`
    also contain `[...]` but never form a standalone bracketed line — so a
    line-anchored match (no DOTALL) discriminates correctly.
    """
    text = predictions_path.read_text(encoding="utf-8", errors="replace")
    items_raw = None
    for line in reversed(text.splitlines()):
        stripped = line.strip()
        m = _PYTHON_LIST_LINE_RE.match(stripped)
        if m:
            items_raw = m.group(1)
            break
    if items_raw is None:
        return [{
            "class_norm": None, "score": None, "raw_class": None,
            "extra_metrics": {}, "missing_in_output": True,
        } for _ in peptides]

    items = [it.strip().strip("'").strip('"') for it in items_raw.split(",") if it.strip()]
    positive_label = parsing.get("positive_label", "ACP")

    records = []
    for i, _ in enumerate(peptides):
        if i >= len(items):
            records.append({
                "class_norm": None, "score": None, "raw_class": None,
                "extra_metrics": {}, "missing_in_output": True,
            })
            continue
        raw = items[i]
        records.append({
            "class_norm": "positive" if raw == positive_label else "negative",
            "score": None,
            "raw_class": raw,
            "extra_metrics": {},
            "missing_in_output": False,
        })
    return records


def _check_score_range(score, out_of_range_counter: dict, tool_id: str):
    """Increment per-tool counter when score lies outside [0,1]. Preserve value."""
    if score is None:
        return
    if score < 0.0 or score > 1.0:
        out_of_range_counter[tool_id] = out_of_range_counter.get(tool_id, 0) + 1


def _compute_agreements(
    matrix: dict[str, dict[str, dict]],
    peptides: list[tuple[str, str]],
    tool_categories: dict[str, str | None],
) -> tuple[list[str], dict[str, dict[str, str]]]:
    """Return (sorted_categories_with_agreement, per_peptide_agreement_map).

    A category appears only if >=2 tools in this run declare it (regardless of
    null vs non-null per peptide; the per-peptide value degrades to single_tool
    when only one of those tools produced a non-null class_norm).
    """
    cat_to_tools: dict[str, list[str]] = {}
    for tid, cat in tool_categories.items():
        if cat is None:
            continue
        cat_to_tools.setdefault(cat, []).append(tid)
    agreement_cats = sorted(c for c, ts in cat_to_tools.items() if len(ts) >= 2)

    per_peptide: dict[str, dict[str, str]] = {}
    for hdr, _seq in peptides:
        flags: dict[str, str] = {}
        for cat in agreement_cats:
            classes = []
            for tid in cat_to_tools[cat]:
                rec = matrix[hdr].get(tid, {})
                cn = rec.get("class_norm")
                if cn in ("positive", "negative"):
                    classes.append(cn)
            if not classes:
                flags[cat] = "no_call"
            elif len(classes) == 1:
                flags[cat] = "single_tool"
            elif all(c == "positive" for c in classes):
                flags[cat] = "consensus_positive"
            elif all(c == "negative" for c in classes):
                flags[cat] = "consensus_negative"
            else:
                flags[cat] = "split"
        per_peptide[hdr] = flags
    return agreement_cats, per_peptide


def _collect_extra_columns(
    matrix: dict[str, dict[str, dict]],
    tools: list[str],
) -> list[tuple[str, str, str]]:
    """Return list of (tool_id, metric_name, unit) for every extra_metric seen."""
    seen: dict[tuple[str, str, str], None] = {}
    for tid in tools:
        for hdr_rec in matrix.values():
            extras = (hdr_rec.get(tid) or {}).get("extra_metrics") or {}
            for name, payload in extras.items():
                seen[(tid, name, payload.get("unit", "none"))] = None
    return list(seen.keys())


def _write_consolidated_csv(
    path: Path,
    peptides: list[tuple[str, str]],
    tools: list[str],
    matrix: dict[str, dict[str, dict]],
    extra_columns: list[tuple[str, str, str]],
    agreement_cats: list[str],
    agreements: dict[str, dict[str, str]],
    cats_in_run: list[str] | None = None,
    cat_aggregates: dict[str, dict[str, dict]] | None = None,
    holistic: dict[str, dict] | None = None,
    structural: dict[str, dict] | None = None,
) -> list[str]:
    """Wide format. Sort: structural_score desc → holistic_score desc.

    Column order: structural_score, structural_max, holistic_score, n_categories_evaluated,
    apex_potency_tag, apex_potency_min_mic_uM, peptide_id, sequence, length, per-tool
    class/score, APEX selectivity (if any), extra_metrics, per-category consensus+mean_score,
    agreement_<cat>.
    """
    cats_in_run = cats_in_run or []
    cat_aggregates = cat_aggregates or {}
    holistic = holistic or {}
    structural = structural or {}

    has_apex_sel = "apex" in tools and any(
        matrix[hdr].get("apex", {}).get("apex_selectivity") for hdr, _ in peptides
    )
    header = [
        "structural_score", "structural_max",
        "holistic_score", "n_categories_evaluated",
        "apex_potency_tag", "apex_potency_min_mic_uM",
        "peptide_id", "sequence", "length",
    ]
    for tid in tools:
        header.append(f"{tid}__class")
        header.append(f"{tid}__score")
    if has_apex_sel:
        header.extend(_APEX_SELECTIVITY_COLS)
    for tid, name, unit in extra_columns:
        header.append(f"{tid}__{name}__{unit}")
    for cat in cats_in_run:
        header.append(f"{cat}__consensus")
        header.append(f"{cat}__mean_score")
    for cat in agreement_cats:
        header.append(f"agreement_{cat}")

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for hdr, seq in peptides:
            h = holistic.get(hdr, {})
            s = structural.get(hdr, {})
            sel_apex = (matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}
            row = [
                s.get("structural_score", 0),
                s.get("structural_max", 0),
                round(h["holistic_score"], 4) if h.get("holistic_score") is not None else None,
                h.get("n_categories_evaluated", 0),
                sel_apex.get("potency_tag"),
                sel_apex.get("potency_min_mic_uM"),
                hdr, seq, len(seq),
            ]
            for tid in tools:
                rec = matrix[hdr].get(tid, {})
                if rec.get("error_batch_failed"):
                    row.append("error_batch_failed")
                else:
                    row.append(rec.get("class_norm"))
                row.append(rec.get("score"))
            if has_apex_sel:
                sel = (matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}
                row.append(sel.get("pathogenic_active"))
                row.append(sel.get("commensal_active"))
                row.append(sel.get("pathogenic_strains_hit"))
                row.append(sel.get("commensal_strains_hit"))
                row.append(sel.get("selectivity_tag"))
            for tid, name, unit in extra_columns:
                rec = matrix[hdr].get(tid, {})
                extras = rec.get("extra_metrics") or {}
                payload = extras.get(name)
                row.append(payload.get("value") if payload else None)
            for cat in cats_in_run:
                agg = cat_aggregates.get(hdr, {}).get(cat) or {}
                row.append(agg.get("consensus"))
                ms = agg.get("mean_score")
                row.append(round(ms, 4) if ms is not None else None)
            for cat in agreement_cats:
                row.append(agreements.get(hdr, {}).get(cat))
            w.writerow(row)
    return header


def _write_consolidated_json(
    path: Path,
    fasta_in: Path,
    peptides: list[tuple[str, str]],
    tools: list[str],
    tool_categories: dict[str, str | None],
    matrix: dict[str, dict[str, dict]],
    agreement_cats: list[str],
    agreements: dict[str, dict[str, str]],
    cats_in_run: list[str] | None = None,
    cat_aggregates: dict[str, dict[str, dict]] | None = None,
    holistic: dict[str, dict] | None = None,
    polarity_map: dict[str, str] | None = None,
    structural: dict[str, dict] | None = None,
) -> None:
    cats_in_run = cats_in_run or []
    cat_aggregates = cat_aggregates or {}
    holistic = holistic or {}
    polarity_map = polarity_map or {}
    structural = structural or {}

    payload = {
        "input_fasta": str(fasta_in),
        "tools": tools,
        "tool_categories": tool_categories,
        "agreement_categories": agreement_cats,
        "categories_in_run": cats_in_run,
        "polarity_map": {c: polarity_map.get(c, "neutral") for c in cats_in_run},
        "ranking": [
            {
                "peptide_id": hdr,
                "structural_score": structural.get(hdr, {}).get("structural_score"),
                "structural_max": structural.get(hdr, {}).get("structural_max"),
                "holistic_score": holistic.get(hdr, {}).get("holistic_score"),
                "apex_potency_tag": ((matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}).get("potency_tag"),
            }
            for hdr, _ in peptides
        ],
        "peptides": [
            {
                "id": hdr,
                "sequence": seq,
                "length": len(seq),
                "structural": structural.get(hdr, {}),
                "holistic": holistic.get(hdr, {}),
                "category_aggregates": cat_aggregates.get(hdr, {}),
                "predictions": {
                    tid: {
                        "class_norm": rec.get("class_norm"),
                        "score": rec.get("score"),
                        "raw_class": rec.get("raw_class"),
                        "extra_metrics": rec.get("extra_metrics") or {},
                        "missing_in_output": rec.get("missing_in_output", True),
                        "error_batch_failed": bool(rec.get("error_batch_failed", False)),
                        **({"apex_selectivity": rec["apex_selectivity"]} if rec.get("apex_selectivity") else {}),
                    }
                    for tid, rec in matrix[hdr].items()
                },
                "agreements": agreements.get(hdr, {}),
            }
            for hdr, seq in peptides
        ],
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _format_extra_value(value, unit: str) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:g} {unit}"
    return f"{value} {unit}"


def _render_report_md(
    out_path: Path,
    fasta_in: Path,
    peptides: list[tuple[str, str]],
    tools: list[str],
    tool_categories: dict[str, str | None],
    matrix: dict[str, dict[str, dict]],
    health: dict[str, dict],
    extra_columns: list[tuple[str, str, str]],
    agreement_cats: list[str],
    agreements: dict[str, dict[str, str]],
    total_seconds: float,
    consolidated_csv: Path,
    consolidated_json: Path,
    cats_in_run: list[str] | None = None,
    cat_aggregates: dict[str, dict[str, dict]] | None = None,
    holistic: dict[str, dict] | None = None,
    structural: dict[str, dict] | None = None,
) -> None:
    cats_in_run = cats_in_run or []
    cat_aggregates = cat_aggregates or {}
    holistic = holistic or {}
    structural = structural or {}
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    lines: list[str] = []
    lines.append("# Audit Report — Pipeline Fase 1")
    lines.append("")
    lines.append(f"- **Input FASTA**: `{fasta_in}`")
    lines.append(f"- **N peptides**: {len(peptides)}")
    lines.append(f"- **Tools ejecutados**: {', '.join(tools)} ({len(tools)})")
    lines.append(f"- **Categorías cubiertas**: "
                 f"{', '.join(sorted({c for c in tool_categories.values() if c}))}")
    lines.append(f"- **Runtime total**: {total_seconds}s")
    lines.append(f"- **Fecha (UTC)**: {now_iso}")
    lines.append("")

    lines.append("## Ranking jerárquico de viabilidad terapéutica")
    lines.append("")
    lines.append("**Sort**: `(structural_score desc, holistic_score desc)`. ")
    lines.append("")
    lines.append("- `structural_score` = suma por categoría según polaridad. Por categoría buena: "
                 "POS=3, SPLIT=2, NEG=1, NONE=0. Por categoría mala: NEG=3, SPLIT=2, POS=1, NONE=0. "
                 "Captura el **perfil estructural** sin promediar magnitudes (un péptido con 4 POS "
                 "buenas y 2 NEG malas siempre va antes que uno con 3 POS buenas y 1 NEG mala).")
    lines.append("- `holistic_score` = `good_mean − bad_mean + APEX_adj + POTENCY_adj`. Desempate "
                 "cuantitativo dentro del mismo `structural_score`.")
    lines.append("- Bonuses APEX: pathogen_specific +0.15, broad_spectrum +0.05, non_active 0, "
                 "commensal_specific −0.20.")
    lines.append("- Bonuses POTENCY: MUY_POTENTE_AMP (min MIC ≤ 5 µM en cualquier cepa) +0.20, "
                 "POTENTE_AMP (min MIC ≤ 10 µM) +0.10. Cuando la cepa con MIC bajo es comensal, "
                 "queda marcada en rojo en el detalle por cepa (alerta visual).")
    lines.append("")
    lines.append("| # | peptide_id | structural | holistic | apex | potency | good_mean | bad_mean |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for i, (hdr, _seq) in enumerate(peptides, 1):
        h = holistic.get(hdr, {})
        s = structural.get(hdr, {})
        hs = h.get("holistic_score")
        hs_str = f"{hs:+.4f}" if isinstance(hs, (int, float)) else "—"
        gm = h.get("good_mean")
        bm = h.get("bad_mean")
        gm_str = f"{gm:.3f}" if isinstance(gm, (int, float)) else "—"
        bm_str = f"{bm:.3f}" if isinstance(bm, (int, float)) else "—"
        tag = h.get("apex_tag") or "—"
        tag_decorated = f"🏆 {tag}" if tag == "pathogen_specific" else tag
        ptag = h.get("potency_tag")
        if ptag == "MUY_POTENTE_AMP":
            ptag_str = "🔥 MUY_POTENTE"
        elif ptag == "POTENTE_AMP":
            ptag_str = "💪 POTENTE"
        else:
            ptag_str = "—"
        ss = s.get("structural_score", 0)
        smax = s.get("structural_max", 0)
        lines.append(f"| {i} | `{hdr}` | **{ss}/{smax}** | **{hs_str}** | {tag_decorated} | "
                     f"{ptag_str} | {gm_str} | {bm_str} |")
    lines.append("")

    lines.append("## Resumen por péptido")
    lines.append("")
    lines.append("| peptide_id | length | n_pos | n_neg | categorias_cubiertas | n_disagreements | apex_tag |")
    lines.append("|---|---|---|---|---|---|---|")
    for hdr, seq in peptides:
        npos = sum(1 for tid in tools if matrix[hdr].get(tid, {}).get("class_norm") == "positive")
        nneg = sum(1 for tid in tools if matrix[hdr].get(tid, {}).get("class_norm") == "negative")
        cats_with_call = sorted({
            tool_categories[tid]
            for tid in tools
            if matrix[hdr].get(tid, {}).get("class_norm") in ("positive", "negative")
            and tool_categories.get(tid)
        })
        ndisagree = sum(1 for cat, flag in agreements.get(hdr, {}).items() if flag == "split")
        sel = (matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}
        tag = sel.get("selectivity_tag")
        tag_str = f"🏆 {tag}" if tag == "pathogen_specific" else (tag or "—")
        lines.append(f"| {hdr} | {len(seq)} | {npos} | {nneg} | "
                     f"{','.join(cats_with_call) or '—'} | {ndisagree} | {tag_str} |")
    lines.append("")
    lines.append("> Detalles por categoría (consensus + mean_score) y métricas continuas: ver "
                 "secciones siguientes.")
    lines.append("")

    lines.append("## Disagreements binarios")
    lines.append("")
    if not agreement_cats:
        lines.append("_Ninguna categoría tiene >=2 tools en este run; no aplica._")
    else:
        any_split = False
        for hdr, _seq in peptides:
            for cat in agreement_cats:
                if agreements.get(hdr, {}).get(cat) != "split":
                    continue
                any_split = True
                detail = []
                for tid in tools:
                    if tool_categories.get(tid) != cat:
                        continue
                    rec = matrix[hdr].get(tid, {})
                    cn = rec.get("class_norm")
                    sc = rec.get("score")
                    sc_str = f"{sc:.3f}" if isinstance(sc, float) else "—"
                    detail.append(f"{tid}={cn}({sc_str})")
                lines.append(f"- **{hdr}** / `{cat}`: {' vs '.join(detail)}")
        if not any_split:
            lines.append("_Ningún disagreement binario detectado en este run._")
    lines.append("")

    lines.append("## Extra metrics")
    lines.append("")
    if not extra_columns:
        lines.append("_No hay tools con extra_metrics activos en este run._")
    else:
        header_cells = ["peptide_id"] + [f"{tid}.{name} ({unit})" for tid, name, unit in extra_columns]
        lines.append("| " + " | ".join(header_cells) + " |")
        lines.append("|" + "|".join(["---"] * len(header_cells)) + "|")
        for hdr, _seq in peptides:
            cells = [hdr]
            for tid, name, unit in extra_columns:
                payload = (matrix[hdr].get(tid, {}).get("extra_metrics") or {}).get(name)
                cells.append(_format_extra_value(payload.get("value") if payload else None, unit))
            lines.append("| " + " | ".join(cells) + " |")
    lines.append("")

    lines.append("## Per-tool health")
    lines.append("")
    lines.append("| tool | category | status | runtime (s) | batches (ok/total) | score_oor | diagnosis |")
    lines.append("|---|---|---|---|---|---|---|")
    for tid in tools:
        h = health.get(tid, {})
        oor = h.get("score_out_of_range", 0)
        diag = h.get("diagnosis") or "—"
        n_b = h.get("n_batches", 0)
        n_b_ok = h.get("n_batches_ok", 0)
        lines.append(
            f"| {tid} | {h.get('category') or '—'} | {h.get('status', '?')} | "
            f"{h.get('runtime_seconds', 0)} | {n_b_ok}/{n_b} | {oor} | {diag} |"
        )
    lines.append("")

    lines.append("## Artefactos")
    lines.append("")
    lines.append(f"- [`consolidated.csv`](./{consolidated_csv.name})")
    lines.append(f"- [`consolidated.json`](./{consolidated_json.name})")
    lines.append("- `tool_health_report.json`")
    lines.append("- `per_tool/<tool_id>/` (raw outputs + run logs)")
    lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


_HTML_STYLE = """
<style>
  *{box-sizing:border-box}
  body{font-family:system-ui,-apple-system,Segoe UI,sans-serif;margin:0;padding:24px;background:#f7f8fa;color:#1a1a1a;line-height:1.4}
  h1,h2,h3{color:#2c3e50;margin-top:0}
  h1{font-size:22px;border-bottom:2px solid #34495e;padding-bottom:8px}
  h2{font-size:18px;margin:0;border-left:4px solid #3498db;padding-left:10px}
  h3{font-size:14px;margin:16px 0 6px}
  .meta{background:#fff;padding:14px 18px;border-radius:6px;border:1px solid #e1e4e8;font-size:13px;margin-bottom:14px}
  .meta b{color:#2c3e50}
  .meta span{margin-right:18px;display:inline-block}
  table{border-collapse:collapse;width:100%;background:#fff;font-size:12px;margin-bottom:8px}
  th,td{border:1px solid #e1e4e8;padding:6px 8px;text-align:left;vertical-align:top}
  thead th{background:#34495e;color:#fff;position:sticky;top:0;z-index:1}
  thead th.sortable{cursor:pointer;user-select:none}
  thead th.sortable:hover{background:#465d75}
  thead th .sort-indicator{font-size:10px;opacity:.55;margin-left:4px}
  thead th.sort-asc .sort-indicator::after{content:"▲";opacity:1}
  thead th.sort-desc .sort-indicator::after{content:"▼";opacity:1}
  thead th.sort-none .sort-indicator::after{content:"⇅";opacity:.45}
  thead th .col-controls{display:block;font-weight:400;font-size:10px;margin-top:3px}
  thead th input.col-filter, thead th select.col-sort{width:100%;padding:1px 3px;font-size:10px;border:1px solid #ccc;border-radius:3px;background:#fff;color:#222}
  tbody tr:hover{background:#eef5fb}
  tbody tr.row-hidden{display:none}
  .chip{display:inline-block;padding:2px 7px;margin:1px 2px;border-radius:10px;font-size:11px;font-weight:500;cursor:default;white-space:nowrap}
  .chip-positive,.chip-pos{background:#d4edda;color:#155724;border:1px solid #b8dcc1}
  .chip-negative,.chip-neg{background:#f8d7da;color:#721c24;border:1px solid #e6b8bd}
  .chip-split{background:#fff3cd;color:#664d03;border:1.5px solid #f0ad4e;font-weight:700}
  .chip-na,.chip-none{background:#e9ecef;color:#6c757d;border:1px dashed #ced4da}
  .pep-id{font-weight:600;color:#2c3e50}
  .badge-pathogen-specific{display:inline-block;background:linear-gradient(180deg,#fde68a,#facc15);color:#5b3a00;border:1px solid #b88a00;font-weight:700;font-size:10px;padding:1px 7px;border-radius:8px;margin-left:5px;letter-spacing:.3px;box-shadow:0 1px 2px rgba(184,138,0,.4)}
  .badge-pathogen-specific::before{content:"🏆 ";font-size:10px}
  .badge-potent{display:inline-block;background:linear-gradient(180deg,#cce5ff,#7eb6ff);color:#003a75;border:1px solid #3b82f6;font-weight:700;font-size:10px;padding:1px 7px;border-radius:8px;margin-left:4px;letter-spacing:.3px;box-shadow:0 1px 2px rgba(59,130,246,.4)}
  .badge-potent::before{content:"💪 ";font-size:10px}
  .badge-very-potent{display:inline-block;background:linear-gradient(180deg,#fecaca,#f97316);color:#5b1900;border:1px solid #c2410c;font-weight:700;font-size:10px;padding:1px 7px;border-radius:8px;margin-left:4px;letter-spacing:.3px;box-shadow:0 1px 2px rgba(194,65,12,.5)}
  .badge-very-potent::before{content:"🔥 ";font-size:10px}
  .strain-pathogen-active{background:#d4edda}
  .strain-commensal-active{background:#f8d7da}
  .strain-ambiguous-active{background:#fff3cd}
  .holistic-cell{font-weight:700;font-family:ui-monospace,monospace}
  .holistic-cell.h-pos{color:#155724}
  .holistic-cell.h-neg{color:#721c24}
  .seq{font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;font-size:12px;background:#f4f6f8;padding:6px 8px;border-radius:4px;word-break:break-all;border:1px solid #e1e4e8}
  details.section{background:#fff;border:1px solid #e1e4e8;border-radius:6px;padding:10px 14px;margin-bottom:12px}
  details.section>summary{cursor:pointer;font-weight:600;color:#2c3e50;padding:4px 0;list-style:none}
  details.section>summary::-webkit-details-marker{display:none}
  details.section>summary::before{content:"▶";display:inline-block;margin-right:8px;font-size:11px;color:#3498db;transition:transform .15s}
  details.section[open]>summary::before{transform:rotate(90deg)}
  details.section[open]>summary{margin-bottom:10px;border-bottom:1px solid #eaecef;padding-bottom:8px}
  details.inline{background:#f9fafb;border:1px solid #e1e4e8;border-radius:4px;padding:6px 10px;margin:4px 0;font-size:12px}
  details.inline>summary{cursor:pointer;color:#3498db;font-weight:500}
  .disagree-box{background:#fff8e1;border-left:4px solid #f0ad4e;padding:12px 16px;border-radius:4px;margin-bottom:12px}
  .disagree-box table{background:#fff8e1}
  .disagree-box thead th{background:#f0ad4e;color:#fff}
  .score-hi{background:#e8f5e9}
  .score-lo{background:#ffebee}
  .status-OK{color:#155724;font-weight:600}
  .status-PROBLEMATIC{color:#721c24;font-weight:600}
  .scroll-x{overflow-x:auto;max-width:100%}
  .toolbar{margin:0 0 8px;padding:8px 12px;background:#eef5fb;border:1px solid #c9dff0;border-radius:4px;font-size:12px;display:flex;flex-wrap:wrap;gap:10px;align-items:center}
  .toolbar button{font-size:11px;padding:3px 10px;border:1px solid #3498db;background:#fff;color:#3498db;border-radius:4px;cursor:pointer}
  .toolbar button:hover{background:#3498db;color:#fff}
  .apex-summary{display:flex;flex-wrap:wrap;gap:8px;margin:8px 0}
  .apex-card{flex:1 1 200px;background:#f4f6f8;border:1px solid #e1e4e8;border-radius:5px;padding:10px 12px;font-size:12px}
  .apex-card .label{color:#6c757d;font-size:11px;text-transform:uppercase;letter-spacing:.4px}
  .apex-card .value{font-size:18px;font-weight:700;color:#2c3e50;margin-top:2px}
  .apex-card .value.unit{font-size:11px;font-weight:400;color:#6c757d;margin-left:3px}
  .footer{margin-top:32px;padding-top:14px;border-top:1px solid #e1e4e8;font-size:12px;color:#6c757d}
  .footer a{color:#3498db;text-decoration:none;margin-right:14px}
  .footer a:hover{text-decoration:underline}
  @media(max-width:700px){body{padding:10px}table{font-size:11px}th,td{padding:4px 5px}}
</style>
""".strip()


_MATRIX_INTERACTIVE_JS = r"""
<script>
(function(){
  const table = document.getElementById('matrix-table');
  if (!table) return;
  const tbody = table.querySelector('tbody');
  const rows = Array.from(tbody.querySelectorAll('tr'));

  const catSortMode = {};
  let currentSort = { col: 'structural_score', dir: 'desc', type: 'num' };

  const CONSENSUS_ORDER = { 'POS': 4, 'SPLIT': 3, 'SINGLE': 3, 'NEG': 2, 'NONE': 1 };
  const TOOL_CLASS_ORDER = { 'positive': 3, 'negative': 1 };

  function getCatCellValue(row, cat, mode){
    const td = row.querySelector('td[data-cat="' + cat + '"]');
    if (!td) return { primary: 0, secondary: 0 };
    if (mode && mode.indexOf('tool:') === 0){
      const tid = mode.slice(5);
      const cls = td.getAttribute('data-tool-' + tid + '-class') || '';
      const sc = parseFloat(td.getAttribute('data-tool-' + tid + '-score'));
      return {
        primary: TOOL_CLASS_ORDER[cls] || 0,
        secondary: isNaN(sc) ? 0 : sc
      };
    }
    const cons = td.getAttribute('data-consensus') || 'NONE';
    const ms = parseFloat(td.getAttribute('data-mean-score'));
    return {
      primary: CONSENSUS_ORDER[cons] || 1,
      secondary: isNaN(ms) ? 0 : ms
    };
  }

  function compareRows(a, b){
    const { col, dir, type } = currentSort;
    let av, bv;
    if (col === 'structural_score'){
      // Sort jerárquico: structural primario, holistic desempate (siempre desc para holistic).
      const sa = +a.dataset.structural, sb = +b.dataset.structural;
      if (sa !== sb) return dir === 'asc' ? sa - sb : sb - sa;
      const ha = +a.dataset.holistic, hb = +b.dataset.holistic;
      return hb - ha;
    }
    if (col === 'rank'){ av = +a.dataset.rank; bv = +b.dataset.rank; }
    else if (col === 'length'){ av = +a.dataset.length; bv = +b.dataset.length; }
    else if (col === 'holistic_score'){ av = +a.dataset.holistic; bv = +b.dataset.holistic; }
    else if (col === 'peptide_id'){ av = a.dataset.pepId; bv = b.dataset.pepId; }
    else if (col.indexOf('cat:') === 0){
      const cat = col.slice(4);
      const mode = catSortMode[cat] || 'consensus';
      const va = getCatCellValue(a, cat, mode);
      const vb = getCatCellValue(b, cat, mode);
      if (va.primary !== vb.primary){
        return dir === 'asc' ? va.primary - vb.primary : vb.primary - va.primary;
      }
      return dir === 'asc' ? va.secondary - vb.secondary : vb.secondary - va.secondary;
    }
    if (type === 'str'){
      const cmp = String(av).localeCompare(String(bv));
      return dir === 'asc' ? cmp : -cmp;
    }
    return dir === 'asc' ? av - bv : bv - av;
  }

  function applySort(){
    const sorted = rows.slice().sort(compareRows);
    sorted.forEach(r => tbody.appendChild(r));
  }

  function updateSortIndicators(){
    table.querySelectorAll('thead th.sortable').forEach(th => {
      th.classList.remove('sort-asc', 'sort-desc', 'sort-none');
      if (th.dataset.col === currentSort.col){
        th.classList.add('sort-' + currentSort.dir);
      } else {
        th.classList.add('sort-none');
      }
    });
  }

  table.querySelectorAll('thead th.sortable').forEach(th => {
    th.addEventListener('click', (e) => {
      if (e.target.tagName === 'SELECT' || e.target.tagName === 'INPUT') return;
      const col = th.dataset.col;
      const type = th.dataset.type || 'str';
      if (currentSort.col === col){
        currentSort.dir = currentSort.dir === 'desc' ? 'asc' : 'desc';
      } else {
        currentSort.col = col;
        currentSort.type = type;
        currentSort.dir = (type === 'num' || type === 'cat') ? 'desc' : 'asc';
      }
      applySort();
      updateSortIndicators();
    });
  });

  table.querySelectorAll('select.col-sort').forEach(sel => {
    sel.addEventListener('click', e => e.stopPropagation());
    sel.addEventListener('change', () => {
      const cat = sel.dataset.cat;
      catSortMode[cat] = sel.value;
      if (currentSort.col === 'cat:' + cat){
        applySort();
        updateSortIndicators();
      }
    });
  });

  function getNumFilter(id){
    const el = document.getElementById(id);
    return (el && el.value !== '') ? parseFloat(el.value) : null;
  }

  function applyFilters(){
    const minStr = getNumFilter('filter-structural-min');
    const minHol = getNumFilter('filter-holistic-min');
    const minLen = getNumFilter('filter-length-min');
    const maxLen = getNumFilter('filter-length-max');
    const onlyPSEl = document.getElementById('filter-only-pathogen-specific');
    const onlyPotentEl = document.getElementById('filter-only-potent-amp');
    const onlyPS = onlyPSEl && onlyPSEl.checked;
    const onlyPotent = onlyPotentEl && onlyPotentEl.checked;
    rows.forEach(r => {
      let visible = true;
      if (minStr !== null && +r.dataset.structural < minStr) visible = false;
      if (minHol !== null && +r.dataset.holistic < minHol) visible = false;
      if (minLen !== null && +r.dataset.length < minLen) visible = false;
      if (maxLen !== null && +r.dataset.length > maxLen) visible = false;
      if (onlyPS && r.dataset.pathogenSpecific !== '1') visible = false;
      if (onlyPotent && r.dataset.potent !== '1') visible = false;
      r.classList.toggle('row-hidden', !visible);
    });
  }

  ['filter-structural-min','filter-holistic-min','filter-length-min','filter-length-max'].forEach(id => {
    const el = document.getElementById(id);
    if (el) el.addEventListener('input', applyFilters);
  });
  const psEl = document.getElementById('filter-only-pathogen-specific');
  const potentEl = document.getElementById('filter-only-potent-amp');
  if (psEl) psEl.addEventListener('change', applyFilters);
  if (potentEl) potentEl.addEventListener('change', applyFilters);
  const resetBtn = document.getElementById('btn-reset-filters');
  if (resetBtn) resetBtn.addEventListener('click', () => {
    ['filter-structural-min','filter-holistic-min','filter-length-min','filter-length-max'].forEach(id => {
      const el = document.getElementById(id); if (el) el.value = '';
    });
    if (psEl) psEl.checked = false;
    if (potentEl) potentEl.checked = false;
    applyFilters();
  });

  updateSortIndicators();
})();
</script>
""".strip()


def _esc(s) -> str:
    return html.escape(str(s)) if s is not None else ""


def _matrix_cell_state(hdr: str, cat: str,
                       tool_categories: dict[str, str | None],
                       matrix: dict[str, dict[str, dict]]) -> tuple[str, str, list]:
    """(state, tooltip, calls) for chip aggregating tools of category `cat` for peptide `hdr`.

    calls = list of (tid, class_norm, score) tuples — exposed so the caller can emit
    per-tool data-* attributes for JS-based sort-by-tool.
    """
    calls = []
    for tid, tcat in tool_categories.items():
        if tcat != cat:
            continue
        rec = matrix[hdr].get(tid, {})
        calls.append((tid, rec.get("class_norm"), rec.get("score")))
    classes = [c for _, c, _ in calls if c in ("positive", "negative")]
    if not classes:
        state = "na"
    elif all(c == "positive" for c in classes):
        state = "positive"
    elif all(c == "negative" for c in classes):
        state = "negative"
    else:
        state = "split"
    parts = []
    for tid, cn, sc in calls:
        sc_str = f" ({sc:.2f})" if isinstance(sc, (int, float)) else ""
        parts.append(f"{tid}: {cn or '—'}{sc_str}")
    return state, "; ".join(parts) if parts else "no tool in this run", calls


def _apex_pathogen_badge(hdr: str, matrix: dict) -> str:
    sel = (matrix.get(hdr, {}).get("apex") or {}).get("apex_selectivity") or {}
    if sel.get("selectivity_tag") == "pathogen_specific":
        return '<span class="badge-pathogen-specific" title="APEX: activo contra patógenos, no contra comensales (selectividad terapéutica ideal)">PATHOGEN SPECIFIC</span>'
    return ""


def _apex_potency_badge(hdr: str, matrix: dict) -> str:
    """Badge POTENTE / MUY POTENTE AMP basado en min(MIC) en cualquier cepa.

    El detail por cepa marca rojo las comensales con MIC ≤ 32, así un POTENTE conseguido
    sólo a costa de una comensal queda visualmente penalizado al expandir detalle.
    """
    sel = (matrix.get(hdr, {}).get("apex") or {}).get("apex_selectivity") or {}
    tag = sel.get("potency_tag")
    if tag is None:
        return ""
    mic = sel.get("potency_min_mic_uM")
    strain = sel.get("potency_strain") or ""
    cat = sel.get("potency_strain_category") or ""
    cat_warning = " ⚠️ vía cepa COMENSAL" if cat == "commensal" else ""
    title = f"min(MIC) = {mic:.2f} µM en {strain} ({cat}){cat_warning}"
    if tag == "MUY_POTENTE_AMP":
        return f'<span class="badge-very-potent" title="{_esc(title)}">MUY POTENTE</span>'
    return f'<span class="badge-potent" title="{_esc(title)}">POTENTE</span>'


def _holistic_css_class(value) -> str:
    if not isinstance(value, (int, float)):
        return ""
    return "h-pos" if value > 0 else ("h-neg" if value < 0 else "")


def _html_meta(fasta_in: Path, peptides, tools, tool_categories, total_seconds, now_iso) -> str:
    cats = sorted({c for c in tool_categories.values() if c})
    return (
        '<div class="meta">'
        f'<span><b>Input</b> <code>{_esc(fasta_in.name)}</code></span>'
        f'<span><b>Péptidos</b> {len(peptides)}</span>'
        f'<span><b>Tools</b> {len(tools)}</span>'
        f'<span><b>Categorías</b> {len(cats)} ({_esc(", ".join(cats))})</span>'
        f'<span><b>Runtime</b> {total_seconds}s</span>'
        f'<span><b>Fecha</b> {_esc(now_iso)}</span>'
        '</div>'
    )


def _html_matrix(peptides, tool_categories, matrix, cats_in_run, cat_aggregates, holistic,
                 structural=None) -> str:
    """Resumen ejecutivo. Tabla principal con sort/filter JS interactivo.

    Cada celda de categoría lleva data-consensus, data-mean-score y data-tool-<tid>-class /
    data-tool-<tid>-score, leídos por JS para sort-by-tool y filtros.

    Sort default = (structural_score desc, holistic_score desc). El JS maneja ambos como
    sort compuesto cuando el usuario clickea el header "structural".
    """
    structural = structural or {}
    cats = list(cats_in_run) or sorted({c for c in tool_categories.values() if c})
    cat_to_tools: dict[str, list[str]] = {}
    for tid, cat in tool_categories.items():
        if cat:
            cat_to_tools.setdefault(cat, []).append(tid)

    parts = [
        '<div class="toolbar">',
        '<span><b>Filtros:</b></span>',
        '<label>structural ≥ <input type="number" step="1" id="filter-structural-min" placeholder="—" style="width:60px"></label>',
        '<label>holistic ≥ <input type="number" step="0.01" id="filter-holistic-min" placeholder="—" style="width:80px"></label>',
        '<label>length ≥ <input type="number" id="filter-length-min" placeholder="—" style="width:60px"></label>',
        '<label>length ≤ <input type="number" id="filter-length-max" placeholder="—" style="width:60px"></label>',
        '<label><input type="checkbox" id="filter-only-pathogen-specific"> solo 🏆 PATHOGEN SPECIFIC</label>',
        '<label><input type="checkbox" id="filter-only-potent-amp"> solo 💪/🔥 POTENTE</label>',
        '<button id="btn-reset-filters" type="button">reset</button>',
        '</div>',
        '<div class="scroll-x"><table id="matrix-table">',
        '<thead><tr>',
        '<th class="sortable sort-none" data-col="rank" data-type="num">#<span class="sort-indicator"></span></th>',
        '<th class="sortable sort-none" data-col="peptide_id" data-type="str">peptide_id<span class="sort-indicator"></span></th>',
        '<th class="sortable sort-none" data-col="length" data-type="num">length<span class="sort-indicator"></span></th>',
        '<th class="sortable sort-desc" data-col="structural_score" data-type="num" '
        'title="Sort jerárquico: structural desc → holistic desc">structural<span class="sort-indicator"></span></th>',
        '<th class="sortable sort-none" data-col="holistic_score" data-type="num">holistic<span class="sort-indicator"></span></th>',
    ]
    for cat in cats:
        tools_in_cat = cat_to_tools.get(cat, [])
        sort_options = '<option value="consensus">consensus + mean score</option>'
        for t in tools_in_cat:
            sort_options += f'<option value="tool:{_esc(t)}">por {_esc(t)}</option>'
        dropdown = (
            f'<span class="col-controls">'
            f'<select class="col-sort" data-cat="{_esc(cat)}">{sort_options}</select>'
            f'</span>'
            if len(tools_in_cat) > 1 else ''
        )
        parts.append(
            f'<th class="sortable sort-none" data-col="cat:{_esc(cat)}" data-type="cat" '
            f'data-cat="{_esc(cat)}" data-tools="{_esc(",".join(tools_in_cat))}">'
            f'{_esc(cat)}<span class="sort-indicator"></span>{dropdown}</th>'
        )
    parts.append('</tr></thead><tbody>')

    for rank, (hdr, seq) in enumerate(peptides, 1):
        h = holistic.get(hdr, {})
        s = structural.get(hdr, {})
        hs = h.get("holistic_score")
        hs_str = f"{hs:+.4f}" if isinstance(hs, (int, float)) else "—"
        hs_cls = _holistic_css_class(hs)
        ss = s.get("structural_score", 0)
        smax = s.get("structural_max", 0)
        sel = (matrix.get(hdr, {}).get("apex") or {}).get("apex_selectivity") or {}
        path_specific = "1" if sel.get("selectivity_tag") == "pathogen_specific" else "0"
        ptag = sel.get("potency_tag") or ""
        is_potent = "1" if ptag in ("POTENTE_AMP", "MUY_POTENTE_AMP") else "0"
        badges = _apex_pathogen_badge(hdr, matrix) + _apex_potency_badge(hdr, matrix)

        parts.append(
            f'<tr data-pep-id="{_esc(hdr)}" data-rank="{rank}" '
            f'data-structural="{ss}" '
            f'data-holistic="{hs if isinstance(hs, (int, float)) else 0}" '
            f'data-length="{len(seq)}" data-pathogen-specific="{path_specific}" '
            f'data-potent="{is_potent}">'
        )
        parts.append(f'<td>{rank}</td>')
        parts.append(f'<td class="pep-id">{_esc(hdr)}{badges}</td>')
        parts.append(f'<td>{len(seq)}</td>')
        parts.append(f'<td style="font-weight:700;color:#2c3e50">{ss}<small style="color:#6c757d">/{smax}</small></td>')
        parts.append(f'<td class="holistic-cell {hs_cls}">{hs_str}</td>')
        for cat in cats:
            state, tip, calls = _matrix_cell_state(hdr, cat, tool_categories, matrix)
            agg = (cat_aggregates.get(hdr, {}).get(cat) or {})
            cons = agg.get("consensus") or "NONE"
            ms = agg.get("mean_score")
            ms_str = f"{ms:.3f}" if isinstance(ms, (int, float)) else "—"
            ms_attr = f"{ms:.6f}" if isinstance(ms, (int, float)) else ""
            label = {"positive": "POS", "negative": "NEG", "split": "SPLIT", "na": "—"}[state]
            tool_attrs = ""
            for tid, cn, sc in calls:
                tool_attrs += f' data-tool-{_esc(tid)}-class="{_esc(cn or "")}"'
                if isinstance(sc, (int, float)):
                    tool_attrs += f' data-tool-{_esc(tid)}-score="{sc}"'
                else:
                    tool_attrs += f' data-tool-{_esc(tid)}-score=""'
            parts.append(
                f'<td data-cat="{_esc(cat)}" data-consensus="{cons}" '
                f'data-mean-score="{ms_attr}"{tool_attrs} '
                f'title="{_esc(tip)}">'
                f'<span class="chip chip-{state}">{label}</span> '
                f'<small style="color:#6c757d">{ms_str}</small>'
                f'</td>'
            )
        parts.append('</tr>')
    parts.append('</tbody></table></div>')
    return "".join(parts)


def _html_disagreements(peptides, tools, tool_categories, matrix,
                        agreement_cats, agreements) -> str:
    rows = []
    for hdr, _ in peptides:
        for cat in agreement_cats:
            if agreements.get(hdr, {}).get(cat) != "split":
                continue
            tool_cells = []
            for tid in tools:
                if tool_categories.get(tid) != cat:
                    continue
                rec = matrix[hdr].get(tid, {})
                cn = rec.get("class_norm")
                sc = rec.get("score")
                sc_str = f"{sc:.3f}" if isinstance(sc, (int, float)) else "—"
                tool_cells.append(f'<b>{_esc(tid)}</b> → {_esc(cn or "—")} ({sc_str})')
            rows.append((hdr, cat, " &nbsp;|&nbsp; ".join(tool_cells)))
    if not rows:
        return '<p><i>Ningún disagreement binario en este run.</i></p>'
    parts = ['<p style="color:#664d03;font-size:12px">Splits entre tools de la misma categoría — '
             'inspeccionar manualmente.</p>',
             '<div class="disagree-box"><table>',
             '<thead><tr><th>peptide_id</th><th>categoría</th><th>tools</th></tr></thead><tbody>']
    for hdr, cat, cells in rows:
        parts.append(f'<tr><td class="pep-id">{_esc(hdr)}</td>'
                     f'<td>{_esc(cat)}</td><td>{cells}</td></tr>')
    parts.append('</tbody></table></div>')
    return "".join(parts)


def _html_drilldown(peptides, tools, tool_categories, matrix) -> str:
    parts = []
    for hdr, seq in peptides:
        badges = _apex_pathogen_badge(hdr, matrix) + _apex_potency_badge(hdr, matrix)
        parts.append(f'<details class="inline"><summary>{_esc(hdr)}{badges} '
                     f'<span style="color:#6c757d;font-weight:400">'
                     f'(length {len(seq)})</span></summary>')
        parts.append(f'<div class="seq">{_esc(seq)}</div>')
        parts.append('<table><thead><tr><th>tool</th><th>category</th>'
                     '<th>class</th><th>score</th><th>extra_metrics</th></tr></thead><tbody>')
        for tid in tools:
            rec = matrix[hdr].get(tid, {})
            cn = rec.get("class_norm")
            sc = rec.get("score")
            extras = rec.get("extra_metrics") or {}
            sc_class = ""
            if isinstance(sc, (int, float)):
                if sc > 0.8: sc_class = ' class="score-hi"'
                elif sc < 0.2: sc_class = ' class="score-lo"'
            sc_str = f"{sc:.3f}" if isinstance(sc, (int, float)) else "—"
            extras_str = "; ".join(
                f"{k}={v.get('value')} {v.get('unit', '')}".strip()
                for k, v in extras.items() if v.get('value') is not None
            ) or "—"
            parts.append(f'<tr{sc_class}><td>{_esc(tid)}</td>'
                         f'<td>{_esc(tool_categories.get(tid) or "—")}</td>'
                         f'<td>{_esc(cn or "—")}</td>'
                         f'<td>{_esc(sc_str)}</td>'
                         f'<td>{_esc(extras_str)}</td></tr>')
        parts.append('</tbody></table></details>')
    return "".join(parts)


def _fmt_mic(v) -> str:
    if not isinstance(v, (int, float)):
        return "—"
    return f"{v:.2f}"


def _load_apex_strain_categories() -> dict[str, str]:
    """Devuelve {metric_name: 'pathogen'|'commensal'|'ambiguous'} desde el YAML."""
    cls_path = REPO_ROOT / "config" / "apex_strain_classification.yaml"
    if not cls_path.exists():
        return {}
    try:
        import yaml
    except ImportError:
        return {}
    cls = yaml.safe_load(cls_path.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    for s in cls.get("pathogenic") or []:
        out[s["metric_name"]] = "pathogen"
    for s in cls.get("commensal") or []:
        out[s["metric_name"]] = "commensal"
    for s in cls.get("ambiguous") or []:
        out[s["metric_name"]] = "ambiguous"
    return out


def _html_apex_block(peptides, matrix) -> str:
    """APEX block dentro de Extra metrics: tag-count summary + tabla por péptido con
    3 medias (pathogen/commensal/total) + <details> inline con detalle por cepa.

    El detalle por cepa colorea cada fila según pathogen (verde si activa) / commensal
    (rojo si activa, indeseable) / ambiguous (amarillo) — facilita ver de un vistazo si
    un POTENTE_AMP se gana matando comensales.
    """
    has_apex = any((matrix[hdr].get("apex") or {}).get("apex_selectivity") for hdr, _ in peptides)
    if not has_apex:
        return ""
    counts = {"pathogen_specific": 0, "commensal_specific": 0,
              "broad_spectrum": 0, "non_active": 0}
    for hdr, _ in peptides:
        tag = ((matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}).get("selectivity_tag")
        if tag in counts:
            counts[tag] += 1
    strain_cat_map = _load_apex_strain_categories()

    parts = [
        '<h3>APEX (multi-cepa MIC, threshold 32 µM)</h3>',
        '<div class="apex-summary">',
        f'<div class="apex-card" style="border-color:#b88a00;background:#fffbe6">'
        f'<div class="label">🏆 pathogen specific</div>'
        f'<div class="value">{counts["pathogen_specific"]}</div></div>',
        f'<div class="apex-card"><div class="label">broad spectrum</div>'
        f'<div class="value">{counts["broad_spectrum"]}</div></div>',
        f'<div class="apex-card"><div class="label">non active</div>'
        f'<div class="value">{counts["non_active"]}</div></div>',
        f'<div class="apex-card" style="border-color:#dc3545;background:#fef2f2">'
        f'<div class="label">⚠️ commensal specific</div>'
        f'<div class="value">{counts["commensal_specific"]}</div></div>',
        '</div>',
        '<p style="font-size:11px;color:#6c757d;margin:6px 0 4px">'
        'Leyenda detalle por cepa: <span class="strain-pathogen-active" style="padding:2px 6px">verde</span> '
        '= patógena activa (deseable); '
        '<span class="strain-commensal-active" style="padding:2px 6px">rojo</span> '
        '= comensal activa (indeseable, daña microbioma); '
        '<span class="strain-ambiguous-active" style="padding:2px 6px">amarillo</span> '
        '= ambigua activa.</p>',
        '<div class="scroll-x"><table>',
        '<thead><tr><th>peptide_id</th><th>tag</th>',
        '<th>MIC media patógenos (µM)</th><th>MIC media comensales (µM)</th>',
        '<th>MIC media total (µM)</th><th>detalle por cepa</th></tr></thead><tbody>',
    ]
    for hdr, _ in peptides:
        rec = matrix[hdr].get("apex") or {}
        sel = rec.get("apex_selectivity") or {}
        extras = rec.get("extra_metrics") or {}
        if not extras:
            continue
        tag = sel.get("selectivity_tag") or "—"
        badges = _apex_pathogen_badge(hdr, matrix) + _apex_potency_badge(hdr, matrix)
        # Per-strain rows (excluir las medias derivadas), coloreadas por categoría
        strain_rows = []
        for k, v in sorted(extras.items()):
            if k.startswith("mean_mic_"):
                continue
            val = v.get("value") if v else None
            unit = v.get("unit", "") if v else ""
            cat = strain_cat_map.get(k, "")
            cls = ""
            if isinstance(val, (int, float)) and val <= 32 and cat:
                cls = f' class="strain-{cat}-active"'
            cat_label = f' <small style="color:#6c757d">({cat})</small>' if cat else ''
            strain_rows.append(
                f'<tr{cls}><td><code>{_esc(k)}</code>{cat_label}</td>'
                f'<td>{_fmt_mic(val)} {_esc(unit)}</td></tr>'
            )
        n_strains = len(strain_rows)
        details_html = (
            f'<details class="inline"><summary>{n_strains} cepas</summary>'
            '<table><thead><tr><th>cepa</th><th>MIC</th></tr></thead><tbody>'
            + "".join(strain_rows) +
            '</tbody></table></details>'
        )
        parts.append(
            f'<tr><td class="pep-id">{_esc(hdr)}{badges}</td>'
            f'<td><code>{_esc(tag)}</code></td>'
            f'<td>{_fmt_mic(sel.get("mean_mic_pathogen_uM"))}</td>'
            f'<td>{_fmt_mic(sel.get("mean_mic_commensal_uM"))}</td>'
            f'<td>{_fmt_mic(sel.get("mean_mic_total_uM"))}</td>'
            f'<td>{details_html}</td></tr>'
        )
    parts.append('</tbody></table></div>')
    return "".join(parts)


def _html_extras(peptides, extra_columns, matrix) -> str:
    has_apex_extras = any(c[0] == "apex" for c in extra_columns)
    other_cols = [c for c in extra_columns if c[0] != "apex"]

    if not has_apex_extras and not other_cols:
        return ''

    parts = []
    if has_apex_extras:
        parts.append(_html_apex_block(peptides, matrix))
    if other_cols:
        parts.append('<h3>Otras métricas continuas (sub-predicciones de tools binarios)</h3>')
        parts.append('<p style="font-size:11px;color:#6c757d;margin:4px 0 8px">'
                     '<b>NO confundir con el score de categoría del Resumen ejecutivo.</b> '
                     'Estos son outputs auxiliares que cada tool emite además de su clase + score '
                     'binario principal. Ejemplo: perseucpp emite <code>prob_cpp</code> (probabilidad '
                     'de ser CPP, alimenta la columna <code>cpp</code>) y por separado <code>prob_eff</code> '
                     '(condicional: si es CPP, ¿es de alta eficiencia?), que aparece aquí como '
                     '<code>efficiency_high_prob</code>. Son métricas distintas del score binario y '
                     'no contribuyen a <code>mean_score</code> ni a <code>holistic_score</code>.</p>')
        parts.append('<div class="scroll-x"><table><thead><tr><th>peptide_id</th>')
        for tid, name, unit in other_cols:
            parts.append(f'<th>{_esc(tid)}<br><small>{_esc(name)} ({_esc(unit)})</small></th>')
        parts.append('</tr></thead><tbody>')
        for hdr, _ in peptides:
            parts.append(f'<tr><td class="pep-id">{_esc(hdr)}</td>')
            for tid, name, _unit in other_cols:
                payload = (matrix[hdr].get(tid, {}).get("extra_metrics") or {}).get(name)
                v = payload.get("value") if payload else None
                cell = f"{v:g}" if isinstance(v, (int, float)) else "—"
                parts.append(f'<td>{_esc(cell)}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table></div>')
    return "".join(parts)


def _html_health(tools, health) -> str:
    parts = ['<table>',
             '<thead><tr><th>tool</th><th>category</th><th>status</th>'
             '<th>runtime (s)</th><th>batches (ok/total)</th>'
             '<th>score_out_of_range</th><th>diagnosis</th>'
             '</tr></thead><tbody>']
    for tid in tools:
        h = health.get(tid, {})
        st = h.get("status", "?")
        n_b = h.get("n_batches", 0)
        n_b_ok = h.get("n_batches_ok", 0)
        parts.append(f'<tr><td>{_esc(tid)}</td>'
                     f'<td>{_esc(h.get("category") or "—")}</td>'
                     f'<td class="status-{_esc(st)}">{_esc(st)}</td>'
                     f'<td>{_esc(h.get("runtime_seconds", 0))}</td>'
                     f'<td>{_esc(n_b_ok)}/{_esc(n_b)}</td>'
                     f'<td>{_esc(h.get("score_out_of_range", 0))}</td>'
                     f'<td>{_esc(h.get("diagnosis") or "—")}</td></tr>')
    parts.append('</tbody></table>')
    return "".join(parts)


def _render_report_html(out_path: Path, fasta_in: Path, peptides, tools,
                        tool_categories, matrix, health, extra_columns,
                        agreement_cats, agreements, total_seconds,
                        consolidated_csv: Path, consolidated_json: Path,
                        report_md: Path,
                        cats_in_run: list[str] | None = None,
                        cat_aggregates: dict | None = None,
                        holistic: dict | None = None,
                        structural: dict | None = None) -> None:
    cats_in_run = cats_in_run or sorted({c for c in tool_categories.values() if c})
    cat_aggregates = cat_aggregates or {}
    holistic = holistic or {}
    structural = structural or {}

    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _section(title: str, body: str, open_default: bool = False) -> str:
        if not body:
            return ""
        attr = " open" if open_default else ""
        return (
            f'<details class="section"{attr}>'
            f'<summary><h2 style="display:inline-block">{_esc(title)}</h2></summary>'
            f'{body}'
            '</details>'
        )

    matrix_html = _html_matrix(peptides, tool_categories, matrix,
                                cats_in_run, cat_aggregates, holistic, structural)
    disagree_html = _html_disagreements(peptides, tools, tool_categories, matrix,
                                         agreement_cats, agreements)
    drilldown_html = _html_drilldown(peptides, tools, tool_categories, matrix)
    extras_html = _html_extras(peptides, extra_columns, matrix)
    health_html = _html_health(tools, health)

    parts = [
        '<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">',
        f'<title>Audit Report — {_esc(fasta_in.name)}</title>',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        _HTML_STYLE,
        '</head><body>',
        f'<h1>Audit Report — Pipeline Fase 1</h1>',
        _html_meta(fasta_in, peptides, tools, tool_categories, total_seconds, now_iso),
        _section("Resumen ejecutivo (matriz interactiva)", matrix_html, open_default=True),
        _section("Disagreements", disagree_html, open_default=False),
        _section("Drill-down por péptido", drilldown_html, open_default=False),
        _section("Extra metrics (APEX + otras)", extras_html, open_default=False),
        _section("Per-tool health", health_html, open_default=False),
        '<div class="footer">Artefactos: ',
        f'<a href="./{consolidated_csv.name}">consolidated.csv</a>',
        f'<a href="./{consolidated_json.name}">consolidated.json</a>',
        f'<a href="./{report_md.name}">REPORT.md</a>',
        '<a href="./tool_health_report.json">tool_health_report.json</a>',
        '</div>',
        _MATRIX_INTERACTIVE_JS,
        '</body></html>',
    ]
    out_path.write_text("".join(parts), encoding="utf-8")


try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


_FILL_POS_TOOL = "FFD4EDDA"
_FILL_NEG_TOOL = "FFF8F9FA"
_FILL_SPLIT = "FFFFF3CD"
_FILL_CONS_POS = "FFC3E6CB"
_FILL_CONS_NEG = "FFE2E3E5"
_FILL_OK = "FFD4EDDA"
_FILL_FAIL = "FFF8D7DA"


def _autosize_and_finalize(ws):
    """Auto-filter on used range + freeze top row + simple width estimation."""
    if ws.max_row <= 1 or ws.max_column < 1:
        ws.freeze_panes = "A2"
        return
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row_idx in range(1, min(ws.max_row, 50) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def _xlsx_matrix_sheet(wb, peptides, tools, matrix, extra_columns,
                       agreement_cats, agreements,
                       cats_in_run=None, cat_aggregates=None, holistic=None,
                       structural=None):
    ws = wb.create_sheet("Matrix")
    cats_in_run = cats_in_run or []
    cat_aggregates = cat_aggregates or {}
    holistic = holistic or {}
    structural = structural or {}
    has_apex_sel = "apex" in tools and any(
        matrix[hdr].get("apex", {}).get("apex_selectivity") for hdr, _ in peptides
    )
    header = [
        "structural_score", "structural_max",
        "holistic_score", "n_categories_evaluated",
        "apex_potency_tag", "apex_potency_min_mic_uM",
        "peptide_id", "sequence", "length",
    ]
    for tid in tools:
        header.append(f"{tid}__class")
        header.append(f"{tid}__score")
    if has_apex_sel:
        header.extend(_APEX_SELECTIVITY_COLS)
    for tid, name, unit in extra_columns:
        header.append(f"{tid}__{name}__{unit}")
    for cat in cats_in_run:
        header.append(f"{cat}__consensus")
        header.append(f"{cat}__mean_score")
    for cat in agreement_cats:
        header.append(f"agreement_{cat}")
    ws.append(header)
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF34495E")
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill

    fill_pos = PatternFill("solid", fgColor=_FILL_POS_TOOL)
    fill_neg = PatternFill("solid", fgColor=_FILL_NEG_TOOL)
    fill_split = PatternFill("solid", fgColor=_FILL_SPLIT)
    fill_cpos = PatternFill("solid", fgColor=_FILL_CONS_POS)
    fill_cneg = PatternFill("solid", fgColor=_FILL_CONS_NEG)

    for hdr, seq in peptides:
        h = holistic.get(hdr, {})
        s = structural.get(hdr, {})
        sel_apex = (matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}
        row = [
            s.get("structural_score", 0),
            s.get("structural_max", 0),
            round(h["holistic_score"], 4) if h.get("holistic_score") is not None else None,
            h.get("n_categories_evaluated", 0),
            sel_apex.get("potency_tag"),
            sel_apex.get("potency_min_mic_uM"),
            hdr, seq, len(seq),
        ]
        for tid in tools:
            rec = matrix[hdr].get(tid, {})
            if rec.get("error_batch_failed"):
                row.append("error_batch_failed")
            else:
                row.append(rec.get("class_norm"))
            row.append(rec.get("score"))
        if has_apex_sel:
            sel = (matrix[hdr].get("apex") or {}).get("apex_selectivity") or {}
            row.append(sel.get("pathogenic_active"))
            row.append(sel.get("commensal_active"))
            row.append(sel.get("pathogenic_strains_hit"))
            row.append(sel.get("commensal_strains_hit"))
            row.append(sel.get("selectivity_tag"))
        for tid, name, unit in extra_columns:
            rec = matrix[hdr].get(tid, {})
            payload = (rec.get("extra_metrics") or {}).get(name)
            row.append(payload.get("value") if payload else None)
        for cat in cats_in_run:
            agg = cat_aggregates.get(hdr, {}).get(cat) or {}
            row.append(agg.get("consensus"))
            ms = agg.get("mean_score")
            row.append(round(ms, 4) if ms is not None else None)
        for cat in agreement_cats:
            row.append(agreements.get(hdr, {}).get(cat))
        ws.append(row)

        r = ws.max_row
        # structural_score (col 1) bold + integer; holistic_score (col 3) bold + float
        ss_cell = ws.cell(row=r, column=1)
        ss_cell.font = Font(bold=True)
        hs_cell = ws.cell(row=r, column=3)
        if isinstance(hs_cell.value, float):
            hs_cell.number_format = "0.0000"
            hs_cell.font = Font(bold=True)
        col_idx = 10  # first tool column (after 9 ranking/header columns)
        for _tid in tools:
            class_cell = ws.cell(row=r, column=col_idx)
            score_cell = ws.cell(row=r, column=col_idx + 1)
            cn = class_cell.value
            if cn == "positive":
                class_cell.fill = fill_pos
                score_cell.fill = fill_pos
            elif cn == "negative":
                class_cell.fill = fill_neg
                score_cell.fill = fill_neg
            if isinstance(score_cell.value, float):
                score_cell.number_format = "0.0000"
            col_idx += 2
        if has_apex_sel:
            col_idx += len(_APEX_SELECTIVITY_COLS)
        col_idx += len(extra_columns)
        # per-category consensus + mean_score (paired)
        for _cat in cats_in_run:
            cons_cell = ws.cell(row=r, column=col_idx)
            mscore_cell = ws.cell(row=r, column=col_idx + 1)
            cv = cons_cell.value
            if cv == "POS":
                cons_cell.fill = fill_cpos
                mscore_cell.fill = fill_cpos
            elif cv == "NEG":
                cons_cell.fill = fill_cneg
                mscore_cell.fill = fill_cneg
            elif cv == "SPLIT":
                cons_cell.fill = fill_split
                cons_cell.font = Font(bold=True)
                mscore_cell.fill = fill_split
            if isinstance(mscore_cell.value, float):
                mscore_cell.number_format = "0.0000"
            col_idx += 2
        for _cat in agreement_cats:
            ag_cell = ws.cell(row=r, column=col_idx)
            v = ag_cell.value
            if v == "split":
                ag_cell.fill = fill_split
                ag_cell.font = Font(bold=True)
            elif v == "consensus_positive":
                ag_cell.fill = fill_cpos
            elif v == "consensus_negative":
                ag_cell.fill = fill_cneg
            col_idx += 1

    _autosize_and_finalize(ws)


def _xlsx_disagreements_sheet(wb, peptides, tools, tool_categories, matrix,
                              agreement_cats, agreements):
    ws = wb.create_sheet("Disagreements")
    header = ["peptide_id", "sequence", "category",
              "tool_A", "class_A", "score_A",
              "tool_B", "class_B", "score_B"]
    ws.append(header)
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FFF0AD4E")
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill

    cat_to_tools: dict[str, list[str]] = {}
    for tid, cat in tool_categories.items():
        if cat is None:
            continue
        cat_to_tools.setdefault(cat, []).append(tid)

    n_rows = 0
    for hdr, seq in peptides:
        for cat in agreement_cats:
            if agreements.get(hdr, {}).get(cat) != "split":
                continue
            tids = cat_to_tools.get(cat, [])
            for i in range(len(tids)):
                for j in range(i + 1, len(tids)):
                    a, b = tids[i], tids[j]
                    ra = matrix[hdr].get(a, {})
                    rb = matrix[hdr].get(b, {})
                    if ra.get("class_norm") == rb.get("class_norm"):
                        continue
                    ws.append([hdr, seq, cat,
                               a, ra.get("class_norm"), ra.get("score"),
                               b, rb.get("class_norm"), rb.get("score")])
                    n_rows += 1

    if n_rows == 0:
        ws.append(["—", "no disagreements in this run", "", "", "", "", "", "", ""])
    _autosize_and_finalize(ws)


def _xlsx_extras_sheet(wb, peptides, extra_columns, matrix):
    ws = wb.create_sheet("Extra_Metrics")
    header = ["peptide_id"] + [f"{tid}__{name}__{unit}" for tid, name, unit in extra_columns]
    ws.append(header)
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF34495E")
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill

    for hdr, _seq in peptides:
        row = [hdr]
        for tid, name, _unit in extra_columns:
            payload = (matrix[hdr].get(tid, {}).get("extra_metrics") or {}).get(name)
            v = payload.get("value") if payload else None
            row.append(v)
        ws.append(row)
        r = ws.max_row
        for col in range(2, len(header) + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    if not extra_columns:
        ws.append(["—"])
    _autosize_and_finalize(ws)


def _xlsx_health_sheet(wb, tools, health):
    ws = wb.create_sheet("Tool_Health")
    header = ["tool_id", "category", "runtime_s", "status",
              "n_batches", "n_batches_ok", "n_batches_failed",
              "score_out_of_range", "diagnosis"]
    ws.append(header)
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF34495E")
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill
    fill_ok = PatternFill("solid", fgColor=_FILL_OK)
    fill_fail = PatternFill("solid", fgColor=_FILL_FAIL)
    for tid in tools:
        h = health.get(tid, {})
        ws.append([tid, h.get("category") or "", h.get("runtime_seconds", 0),
                   h.get("status", "?"),
                   h.get("n_batches", 0), h.get("n_batches_ok", 0), h.get("n_batches_failed", 0),
                   h.get("score_out_of_range", 0),
                   h.get("diagnosis") or ""])
        r = ws.max_row
        st_cell = ws.cell(row=r, column=4)
        f = fill_ok if st_cell.value == "OK" else fill_fail
        for col in range(1, len(header) + 1):
            ws.cell(row=r, column=col).fill = f
    _autosize_and_finalize(ws)


def _xlsx_run_info_sheet(wb, fasta_in, peptides, tools, total_seconds, health):
    ws = wb.create_sheet("Run_Info")
    ws.append(["key", "value"])
    bold = Font(bold=True, color="FFFFFFFF")
    head_fill = PatternFill("solid", fgColor="FF34495E")
    for c in ws[1]:
        c.font = bold
        c.fill = head_fill
    n_ok = sum(1 for h in health.values() if h.get("status") == "OK")
    rows = [
        ("input_file", str(fasta_in)),
        ("n_peptides", len(peptides)),
        ("n_tools", len(tools)),
        ("n_tools_ok", n_ok),
        ("runtime_total_s", total_seconds),
        ("datetime_iso", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("tools_executed", ", ".join(tools)),
    ]
    for k, v in rows:
        ws.append([k, v])
        kc = ws.cell(row=ws.max_row, column=1)
        kc.font = Font(bold=True)
    _autosize_and_finalize(ws)


def _write_consolidated_xlsx(path: Path, fasta_in: Path, peptides, tools,
                             tool_categories, matrix, extra_columns,
                             agreement_cats, agreements, health,
                             total_seconds,
                             cats_in_run=None, cat_aggregates=None, holistic=None,
                             structural=None) -> None:
    if not _OPENPYXL_OK:
        return
    wb = Workbook()
    wb.remove(wb.active)
    _xlsx_matrix_sheet(wb, peptides, tools, matrix, extra_columns,
                       agreement_cats, agreements,
                       cats_in_run=cats_in_run, cat_aggregates=cat_aggregates,
                       holistic=holistic, structural=structural)
    _xlsx_disagreements_sheet(wb, peptides, tools, tool_categories, matrix,
                              agreement_cats, agreements)
    _xlsx_extras_sheet(wb, peptides, extra_columns, matrix)
    _xlsx_health_sheet(wb, tools, health)
    _xlsx_run_info_sheet(wb, fasta_in, peptides, tools, total_seconds, health)
    wb.save(str(path))


def _write_fasta(path: Path, peptides: list[tuple[str, str]]) -> None:
    """Write peptides to FASTA in input order (no validation; assumes already-clean upper AAs)."""
    with path.open("w", encoding="utf-8") as f:
        for hdr, seq in peptides:
            f.write(f">{hdr}\n{seq}\n")


def _failed_batch_records(n: int) -> list[dict]:
    """Records for peptides whose batch crashed/timed-out — class None, raw 'error_batch_failed'."""
    return [{
        "class_norm": None, "score": None,
        "raw_class": "error_batch_failed",
        "extra_metrics": {}, "missing_in_output": True,
        "error_batch_failed": True,
    } for _ in range(n)]


def _merge_batch_outputs(batch_paths: list[Path], canonical: Path, output_format: str) -> None:
    """Concatenate per-batch raw outputs into a single canonical file (best-effort, for inspection).

    For CSV: keeps the header from the first batch and skips it for subsequent batches.
    For other formats (e.g. stdout-captured text): plain concatenation.
    The orchestrator parses each batch separately, so this file is purely for the user.
    """
    if not batch_paths:
        canonical.write_text("", encoding="utf-8")
        return
    if output_format == "csv":
        with canonical.open("w", encoding="utf-8") as out:
            for i, p in enumerate(batch_paths):
                if not p.exists():
                    continue
                txt = p.read_text(encoding="utf-8", errors="replace")
                if not txt:
                    continue
                if i == 0:
                    out.write(txt if txt.endswith("\n") else txt + "\n")
                else:
                    lines = txt.splitlines(True)
                    if lines:
                        out.write("".join(lines[1:]))
                        if lines[-1] and not lines[-1].endswith("\n"):
                            out.write("\n")
    else:
        with canonical.open("w", encoding="utf-8") as out:
            for p in batch_paths:
                if p.exists():
                    out.write(p.read_text(encoding="utf-8", errors="replace"))
                    out.write("\n")


def _run_tool_batched(
    tool_id: str,
    peptides: list[tuple[str, str]],
    tool_dir: Path,
    config_path: str,
    batch_size: int,
    parsing: dict,
    output_format: str,
) -> tuple[list[dict], dict, list[Path]]:
    """Execute tool over sequential batches; merge raw outputs; return per-peptide records.

    Failed batches (timeout, crash, parse_error) do NOT abort the run — peptides of that batch
    receive `error_batch_failed` records and the run continues with the next batch.

    Returns (records_in_input_order, health_summary, temp_fasta_paths_to_cleanup).
    """
    n_pep = len(peptides)
    n_batches = max(1, math.ceil(n_pep / batch_size))

    records_all: list[dict] = []
    runtime_total = 0.0
    n_ok_batches = 0
    n_failed_batches = 0
    diagnoses: list[str] = []
    last_stderr_tail: str | None = None
    raw_batch_paths: list[Path] = []
    temp_fastas: list[Path] = []

    batches_dir = tool_dir / "_batches"
    batches_dir.mkdir(exist_ok=True)

    for b in range(n_batches):
        start = b * batch_size
        end = min(start + batch_size, n_pep)
        batch_peps = peptides[start:end]

        batch_dir = batches_dir / f"batch_{b:03d}"
        batch_dir.mkdir(exist_ok=True)
        batch_fasta = batch_dir / f"input_{tool_id}_batch_{b:03d}.fasta"
        _write_fasta(batch_fasta, batch_peps)
        temp_fastas.append(batch_fasta)

        print(f"  [{tool_id}] batch {b+1}/{n_batches} (peptides {start+1}-{end})")

        result = run_tool(
            tool_id=tool_id,
            peptides_fasta=batch_fasta,
            output_dir=batch_dir,
            pipeline_config_path=config_path,
        )
        runtime_total += result.runtime_seconds
        if result.stderr_tail:
            last_stderr_tail = result.stderr_tail

        if result.status == "OK" and result.predictions_path:
            raw_batch_paths.append(result.predictions_path)
            try:
                batch_records = _parse_predictions(tool_id, result.predictions_path, parsing, batch_peps)
                records_all.extend(batch_records)
                n_ok_batches += 1
            except Exception as exc:
                diagnoses.append(f"batch_{b}: parse_error: {exc.__class__.__name__}: {exc}")
                records_all.extend(_failed_batch_records(len(batch_peps)))
                n_failed_batches += 1
        else:
            diagnoses.append(f"batch_{b}: {result.diagnosis}")
            records_all.extend(_failed_batch_records(len(batch_peps)))
            n_failed_batches += 1

    canonical = tool_dir / f"predictions_{tool_id}.{output_format}"
    if raw_batch_paths:
        _merge_batch_outputs(raw_batch_paths, canonical, output_format)

    if n_failed_batches == 0:
        final_status = "OK"
        final_diagnosis: str | None = None
    elif n_ok_batches == 0:
        final_status = "PROBLEMATIC"
        final_diagnosis = "all batches failed: " + " | ".join(diagnoses)
    else:
        final_status = "OK"
        final_diagnosis = f"partial: {n_failed_batches}/{n_batches} batches failed: " + " | ".join(diagnoses)

    health = {
        "status": final_status,
        "diagnosis": final_diagnosis,
        "stderr_tail": last_stderr_tail,
        "runtime_seconds": round(runtime_total, 2),
        "n_batches": n_batches,
        "n_batches_ok": n_ok_batches,
        "n_batches_failed": n_failed_batches,
        "predictions_path": str(canonical) if raw_batch_paths else None,
    }
    return records_all, health, temp_fastas


_APEX_SELECTIVITY_COLS = [
    "apex__pathogenic_active",
    "apex__commensal_active",
    "apex__pathogenic_strains_hit__count",
    "apex__commensal_strains_hit__count",
    "apex__selectivity_tag",
]


def _load_apex_classification(repo_root: Path) -> dict | None:
    cfg_path = repo_root / "config" / "apex_strain_classification.yaml"
    if not cfg_path.exists():
        return None
    try:
        import yaml
    except ImportError:
        return None
    return yaml.safe_load(cfg_path.read_text(encoding="utf-8"))


def _apply_apex_selectivity(records: list[dict], classification: dict | None) -> None:
    """Post-process APEX records → derive selectivity tag + 3 mean-MIC summary metrics.

    Threshold (MIC ≤ T → active) configurable via classification['threshold']['active_mic_uM'].
    Ambiguous strains are excluded from the aggregate but preserved as raw extra_metric columns.

    APEX class_norm permanece `None` (extra_only). Decisión 2026-05-01: APEX no entra en el
    eje binario class_norm porque el threshold 32 µM es subjetivo y la categoría antimicrobial
    se sostiene con `antibp3` como único proveedor binario. APEX aporta selectivity_tag (descriptor
    biológico crítico para "pathogen specific") + 3 medias de MIC (pathogen/commensal/total).
    """
    if not classification:
        return
    threshold = float((classification.get("threshold") or {}).get("active_mic_uM", 32.0))
    pathogenic_metrics = [s["metric_name"] for s in classification.get("pathogenic") or []]
    commensal_metrics = [s["metric_name"] for s in classification.get("commensal") or []]

    # Build strain → category lookup for badge metadata
    strain_category = {}
    for s in classification.get("pathogenic") or []:
        strain_category[s["metric_name"]] = "pathogen"
    for s in classification.get("commensal") or []:
        strain_category[s["metric_name"]] = "commensal"
    for s in classification.get("ambiguous") or []:
        strain_category[s["metric_name"]] = "ambiguous"

    for rec in records:
        extras = rec.get("extra_metrics") or {}
        if not extras:
            continue
        path_vals = [
            p["value"] for m in pathogenic_metrics
            if (p := extras.get(m)) and p.get("value") is not None
        ]
        comm_vals = [
            p["value"] for m in commensal_metrics
            if (p := extras.get(m)) and p.get("value") is not None
        ]
        path_hits = sum(1 for v in path_vals if v <= threshold)
        comm_hits = sum(1 for v in comm_vals if v <= threshold)
        path_active = 1 if path_hits > 0 else 0
        comm_active = 1 if comm_hits > 0 else 0
        if path_active == 1 and comm_active == 0:
            tag = "pathogen_specific"
        elif path_active == 0 and comm_active == 1:
            tag = "commensal_specific"
        elif path_active == 1 and comm_active == 1:
            tag = "broad_spectrum"
        else:
            tag = "non_active"
        all_vals = path_vals + comm_vals
        mean_path = sum(path_vals) / len(path_vals) if path_vals else None
        mean_comm = sum(comm_vals) / len(comm_vals) if comm_vals else None
        mean_total = sum(all_vals) / len(all_vals) if all_vals else None

        # Potency badge — min MIC en CUALQUIER cepa clasificada (incluye comensales).
        # El detail por cepa marca rojo las comensales con MIC ≤ 32, así que un POTENTE_AMP
        # ganado solo por cepa comensal queda visualmente penalizado en el HTML.
        candidate_strains = []  # (mic, strain_name, category)
        for m, p in extras.items():
            if m.startswith("mean_mic_") or not isinstance(p, dict):
                continue
            v = p.get("value")
            cat = strain_category.get(m)
            if isinstance(v, (int, float)) and cat in ("pathogen", "commensal"):
                candidate_strains.append((v, m, cat))
        potency_tag = None
        potency_strain = None
        if candidate_strains:
            min_mic, min_strain, min_cat = min(candidate_strains, key=lambda t: t[0])
            if min_mic <= _POTENCY_THRESHOLDS_uM["MUY_POTENTE_AMP"]:
                potency_tag = "MUY_POTENTE_AMP"
                potency_strain = (min_mic, min_strain, min_cat)
            elif min_mic <= _POTENCY_THRESHOLDS_uM["POTENTE_AMP"]:
                potency_tag = "POTENTE_AMP"
                potency_strain = (min_mic, min_strain, min_cat)

        rec["apex_selectivity"] = {
            "pathogenic_active": path_active,
            "commensal_active": comm_active,
            "pathogenic_strains_hit": path_hits,
            "commensal_strains_hit": comm_hits,
            "selectivity_tag": tag,
            "mean_mic_pathogen_uM": mean_path,
            "mean_mic_commensal_uM": mean_comm,
            "mean_mic_total_uM": mean_total,
            "potency_tag": potency_tag,
            "potency_min_mic_uM": potency_strain[0] if potency_strain else None,
            "potency_strain": potency_strain[1] if potency_strain else None,
            "potency_strain_category": potency_strain[2] if potency_strain else None,
        }
        extras["mean_mic_pathogen"] = {"value": mean_path, "unit": "uM"}
        extras["mean_mic_commensal"] = {"value": mean_comm, "unit": "uM"}
        extras["mean_mic_total"] = {"value": mean_total, "unit": "uM"}


def _load_categories_config(repo_root: Path) -> dict | None:
    cfg_path = repo_root / "config" / "categories_config.yaml"
    if not cfg_path.exists():
        return None
    try:
        import yaml
    except ImportError:
        return None
    return yaml.safe_load(cfg_path.read_text(encoding="utf-8"))


def _polarity_map(cats_cfg: dict | None) -> dict[str, str]:
    """Return {category_name: polarity}. Categories without explicit polarity → 'neutral'."""
    if not cats_cfg:
        return {}
    cats = cats_cfg.get("categories") or {}
    return {k: (v.get("polarity") or "neutral") for k, v in cats.items()}


_APEX_HOLISTIC_ADJUST = {
    "pathogen_specific": 0.15,
    "broad_spectrum": 0.05,
    "non_active": 0.0,
    "commensal_specific": -0.20,
}

# Potencia AMP (independiente del selectivity_tag, basada en min MIC en CUALQUIER cepa).
# Bonus excluyente: si MIN_MIC ≤ 5 → MUY_POTENTE (+0.20), elif ≤ 10 → POTENTE (+0.10), else 0.
# El detail por cepa marca visualmente cuándo esa "potencia" está sobre una comensal (rojo).
_POTENCY_THRESHOLDS_uM = {"MUY_POTENTE_AMP": 5.0, "POTENTE_AMP": 10.0}
_POTENCY_HOLISTIC_ADJUST = {"MUY_POTENTE_AMP": 0.20, "POTENTE_AMP": 0.10}


def _compute_category_aggregates(
    matrix: dict[str, dict[str, dict]],
    peptides: list[tuple[str, str]],
    tool_categories: dict[str, str | None],
) -> tuple[list[str], dict[str, dict[str, dict]]]:
    """Per peptide × category: consensus + mean_score across binary tools of that category.

    Returns (sorted_cats_in_run, {hdr: {cat: {consensus, mean_score, n_tools}}}).

    `consensus`: POS (all tools positive, or single-tool positive), NEG (all negative or
    single-tool negative), SPLIT (mixed), NONE (no binary call from any tool).
    `mean_score`: mean of non-None scores from binary tools in that category. None if no scores.
    """
    cat_to_tools: dict[str, list[str]] = {}
    for tid, cat in tool_categories.items():
        if cat is None:
            continue
        cat_to_tools.setdefault(cat, []).append(tid)
    cats = sorted(cat_to_tools.keys())

    out: dict[str, dict[str, dict]] = {}
    for hdr, _ in peptides:
        per_cat: dict[str, dict] = {}
        for cat in cats:
            tids = cat_to_tools[cat]
            classes: list[str] = []
            scores: list[float] = []
            for tid in tids:
                rec = matrix[hdr].get(tid, {})
                cn = rec.get("class_norm")
                sc = rec.get("score")
                if cn in ("positive", "negative"):
                    classes.append(cn)
                if isinstance(sc, (int, float)):
                    scores.append(float(sc))
            if not classes:
                consensus = "NONE"
            elif all(c == "positive" for c in classes):
                consensus = "POS"
            elif all(c == "negative" for c in classes):
                consensus = "NEG"
            else:
                consensus = "SPLIT"
            mean_score = sum(scores) / len(scores) if scores else None
            per_cat[cat] = {"consensus": consensus, "mean_score": mean_score, "n_tools": len(tids)}
        out[hdr] = per_cat
    return cats, out


def _compute_holistic_scores(
    peptides: list[tuple[str, str]],
    cat_aggregates: dict[str, dict[str, dict]],
    polarity_map: dict[str, str],
    matrix: dict[str, dict[str, dict]],
) -> dict[str, dict]:
    """Per peptide → {holistic_score, n_categories_evaluated, good_mean, bad_mean,
                       apex_adjustment, potency_adjustment, apex_tag, potency_tag}.

    Formula:
        holistic_score = good_mean − bad_mean + apex_adjustment + potency_adjustment

    - apex_adjustment ∈ {+0.15 path_specific, +0.05 broad, 0 non, −0.20 commensal}
    - potency_adjustment ∈ {+0.20 MUY_POTENTE_AMP, +0.10 POTENTE_AMP, 0} — basado en
      min(MIC) sobre cualquier cepa clasificada (pathogen o commensal). Es independiente
      de selectivity_tag, así que un péptido puede acumular ambos bonuses.

    Missing categories excluidas del promedio (no cuentan como 0). `n_categories_evaluated`
    para transparencia.
    """
    out: dict[str, dict] = {}
    for hdr, _ in peptides:
        per_cat = cat_aggregates.get(hdr, {})
        good_scores: list[float] = []
        bad_scores: list[float] = []
        n_evaluated = 0
        for cat, agg in per_cat.items():
            ms = agg.get("mean_score")
            if ms is None:
                continue
            n_evaluated += 1
            pol = polarity_map.get(cat, "neutral")
            if pol == "good":
                good_scores.append(ms)
            elif pol == "bad":
                bad_scores.append(ms)
        good_mean = sum(good_scores) / len(good_scores) if good_scores else None
        bad_mean = sum(bad_scores) / len(bad_scores) if bad_scores else None

        apex_rec = matrix[hdr].get("apex", {}) or {}
        apex_sel = apex_rec.get("apex_selectivity") or {}
        apex_tag = apex_sel.get("selectivity_tag")
        apex_adj = _APEX_HOLISTIC_ADJUST.get(apex_tag, 0.0)
        potency_tag = apex_sel.get("potency_tag")
        potency_adj = _POTENCY_HOLISTIC_ADJUST.get(potency_tag, 0.0)

        holistic = (good_mean or 0.0) - (bad_mean or 0.0) + apex_adj + potency_adj
        out[hdr] = {
            "holistic_score": holistic,
            "n_categories_evaluated": n_evaluated,
            "good_mean": good_mean,
            "bad_mean": bad_mean,
            "apex_adjustment": apex_adj,
            "apex_tag": apex_tag,
            "potency_adjustment": potency_adj,
            "potency_tag": potency_tag,
        }
    return out


# Per-category tier score por polaridad. Captura "nivel de bondad" estructural sin promediar.
_TIER_GOOD = {"POS": 3, "SPLIT": 2, "NEG": 1, "NONE": 0}
_TIER_BAD = {"NEG": 3, "SPLIT": 2, "POS": 1, "NONE": 0}


def _compute_structural_scores(
    peptides: list[tuple[str, str]],
    cat_aggregates: dict[str, dict[str, dict]],
    polarity_map: dict[str, str],
) -> dict[str, dict]:
    """Per peptide → {structural_score, structural_max, n_pos_good, n_neg_bad, ...}.

    structural_score: suma de tier scores por categoría según polarity. Para cada categoría:
      - polarity=good: POS=3, SPLIT=2, NEG=1, NONE=0
      - polarity=bad : NEG=3, SPLIT=2, POS=1, NONE=0
      - polarity=neutral: 0 (no contribuye)

    Sirve como **primer nivel** de ordenación del ranking: todos los péptidos con perfil
    estructural mejor (más POS en good, más NEG en bad, SPLITs como intermedio) van primero,
    independientemente de la magnitud de los scores. holistic_score es desempate dentro del
    mismo tier.

    structural_max devuelto para transparencia (= 3 × n_categorías good+bad evaluadas).
    """
    out: dict[str, dict] = {}
    for hdr, _ in peptides:
        per_cat = cat_aggregates.get(hdr, {})
        score = 0
        max_possible = 0
        n_pos_good = n_split_good = n_neg_good = 0
        n_neg_bad = n_split_bad = n_pos_bad = 0
        for cat, agg in per_cat.items():
            pol = polarity_map.get(cat, "neutral")
            cons = agg.get("consensus") or "NONE"
            if pol == "good":
                score += _TIER_GOOD.get(cons, 0)
                max_possible += 3
                if cons == "POS": n_pos_good += 1
                elif cons == "SPLIT": n_split_good += 1
                elif cons == "NEG": n_neg_good += 1
            elif pol == "bad":
                score += _TIER_BAD.get(cons, 0)
                max_possible += 3
                if cons == "NEG": n_neg_bad += 1
                elif cons == "SPLIT": n_split_bad += 1
                elif cons == "POS": n_pos_bad += 1
            # neutral: no contribuye
        out[hdr] = {
            "structural_score": score,
            "structural_max": max_possible,
            "n_pos_good": n_pos_good,
            "n_split_good": n_split_good,
            "n_neg_good": n_neg_good,
            "n_neg_bad": n_neg_bad,
            "n_split_bad": n_split_bad,
            "n_pos_bad": n_pos_bad,
        }
    return out


def _resolve_input(value: str) -> Path:
    """Resolve --input: absolute → literal; with '/' → cwd-relative; bare → INPUTS_DIR."""
    p = Path(value)
    if p.is_absolute():
        if p.exists():
            return p.resolve()
        sys.exit(f"Input file not found. Tried: {p}")
    if "/" in value or "\\" in value:
        cand = (Path.cwd() / p).resolve()
        if cand.exists():
            return cand
        sys.exit(f"Input file not found. Tried: {cand}")
    cand_inputs = (INPUTS_DIR / p).resolve()
    if cand_inputs.exists():
        return cand_inputs
    cand_cwd = (Path.cwd() / p).resolve()
    if cand_cwd.exists():
        return cand_cwd
    sys.exit(f"Input file not found. Tried: {cand_inputs}, {cand_cwd}")


def _resolve_output(explicit: Path | None, input_path: Path) -> Path:
    """Resolve --output: explicit literal, else OUTPUTS_DIR/<stem>_<ISO_ts>/ with collision suffix."""
    if explicit is not None:
        return explicit.resolve()
    stem = input_path.stem
    ts = datetime.now().strftime("%Y-%m-%dT%H%M")
    base = OUTPUTS_DIR / f"{stem}_{ts}"
    if not base.exists():
        return base
    n = 2
    while True:
        cand = OUTPUTS_DIR / f"{stem}_{ts}_{n}"
        if not cand.exists():
            return cand
        n += 1


def main():
    ap = argparse.ArgumentParser(description="Pipeline Fase 1 orchestrator (E2E)")
    ap.add_argument("--input", required=True,
                    help="FASTA: bare name (looked up in Inputs/), relative path, or absolute path")
    ap.add_argument("--output", default=None, type=Path,
                    help="Output dir (default: Outputs/<input_stem>_<ISO_ts>/)")
    ap.add_argument("--tools", default="all",
                    help="Comma-separated tool IDs, or 'all' for the 10 OK tools")
    ap.add_argument("--config", default=str(REPO_ROOT / "config" / "pipeline_config.yaml"))
    ap.add_argument("--batch-size", type=int, default=100,
                    help="Peptides per batch (default 100, min 1). Per-tool override via "
                         "`batch_size_override` in pipeline_config.yaml.")
    args = ap.parse_args()
    batch_size_global = max(1, args.batch_size)

    fasta_in: Path = _resolve_input(args.input)
    out_root: Path = _resolve_output(args.output, fasta_in)
    out_root.mkdir(parents=True, exist_ok=True)
    per_tool_root = out_root / "per_tool"
    per_tool_root.mkdir(exist_ok=True)

    tools = DEFAULT_TOOLS if args.tools == "all" else [t.strip() for t in args.tools.split(",") if t.strip()]
    peptides = parse_fasta(fasta_in)
    if not peptides:
        sys.exit(f"No peptides parsed from {fasta_in}")

    cfg = load_pipeline_config(args.config)

    health: dict[str, dict] = {}
    matrix: dict[str, dict[str, dict]] = {hdr: {} for hdr, _ in peptides}
    tool_categories: dict[str, str | None] = {}
    score_oor: dict[str, int] = {}
    temp_fastas_to_cleanup: list[Path] = []
    pipeline_t0 = time.monotonic()

    for tool_id in tools:
        tool_dir = per_tool_root / tool_id
        tool_dir.mkdir(exist_ok=True)
        try:
            tool_cfg = get_tool_config(tool_id, cfg)
        except KeyError:
            health[tool_id] = {
                "status": "PROBLEMATIC", "diagnosis": "unknown_tool",
                "runtime_seconds": 0.0, "category": None, "score_out_of_range": 0,
                "n_batches": 0, "n_batches_ok": 0, "n_batches_failed": 0,
            }
            tool_categories[tool_id] = None
            for hdr, _ in peptides:
                matrix[hdr][tool_id] = {
                    "class_norm": None, "score": None, "raw_class": None,
                    "extra_metrics": {}, "missing_in_output": True,
                }
            continue

        category = tool_cfg.get("category")
        tool_categories[tool_id] = category

        run_cmd = tool_cfg.get("run_command") or {}
        parsing = run_cmd.get("output_parsing") or {}
        output_format = run_cmd.get("output_format", "csv")
        tool_batch_size = max(1, int(tool_cfg.get("batch_size_override") or batch_size_global))

        n_pep = len(peptides)
        n_batches_planned = max(1, math.ceil(n_pep / tool_batch_size))
        print(f"[{tool_id}] running {n_pep} peptides in {n_batches_planned} batch(es) "
              f"of up to {tool_batch_size}")

        records, batch_health, temp_fastas = _run_tool_batched(
            tool_id=tool_id,
            peptides=peptides,
            tool_dir=tool_dir,
            config_path=args.config,
            batch_size=tool_batch_size,
            parsing=parsing,
            output_format=output_format,
        )
        temp_fastas_to_cleanup.extend(temp_fastas)

        if tool_id == "apex":
            apex_cls = _load_apex_classification(REPO_ROOT)
            _apply_apex_selectivity(records, apex_cls)

        health[tool_id] = {
            **batch_health,
            "category": category,
            "score_out_of_range": 0,
        }

        for (hdr, _seq), rec in zip(peptides, records):
            matrix[hdr][tool_id] = rec
            _check_score_range(rec.get("score"), score_oor, tool_id)
        health[tool_id]["score_out_of_range"] = score_oor.get(tool_id, 0)

    for p in temp_fastas_to_cleanup:
        try:
            p.unlink(missing_ok=True)
        except OSError:
            pass

    total_seconds = round(time.monotonic() - pipeline_t0, 2)

    agreement_cats, agreements = _compute_agreements(matrix, peptides, tool_categories)
    extra_columns = _collect_extra_columns(matrix, tools)

    cats_cfg = _load_categories_config(REPO_ROOT)
    polarity_map = _polarity_map(cats_cfg)
    cats_in_run, cat_aggregates = _compute_category_aggregates(matrix, peptides, tool_categories)
    holistic = _compute_holistic_scores(peptides, cat_aggregates, polarity_map, matrix)
    structural = _compute_structural_scores(peptides, cat_aggregates, polarity_map)

    # Ranking jerárquico:
    #   1) structural_score desc (perfil POS/NEG/SPLIT por categoría según polaridad)
    #   2) holistic_score desc (desempate cuantitativo dentro del mismo tier)
    peptides = sorted(
        peptides,
        key=lambda hs: (
            structural.get(hs[0], {}).get("structural_score", 0),
            holistic.get(hs[0], {}).get("holistic_score", 0.0),
        ),
        reverse=True,
    )

    consolidated_csv = out_root / "consolidated.csv"
    _write_consolidated_csv(
        consolidated_csv, peptides, tools, matrix, extra_columns, agreement_cats, agreements,
        cats_in_run=cats_in_run, cat_aggregates=cat_aggregates, holistic=holistic,
        structural=structural,
    )

    consolidated_json = out_root / "consolidated.json"
    _write_consolidated_json(
        consolidated_json, fasta_in, peptides, tools, tool_categories,
        matrix, agreement_cats, agreements,
        cats_in_run=cats_in_run, cat_aggregates=cat_aggregates, holistic=holistic,
        polarity_map=polarity_map, structural=structural,
    )

    health_report = out_root / "tool_health_report.json"
    health_report.write_text(json.dumps({
        "input_fasta": str(fasta_in),
        "n_peptides": len(peptides),
        "n_tools_requested": len(tools),
        "n_tools_ok": sum(1 for h in health.values() if h["status"] == "OK"),
        "total_seconds": total_seconds,
        "tools": health,
    }, indent=2), encoding="utf-8")

    report_md = out_root / "REPORT.md"
    _render_report_md(
        report_md, fasta_in, peptides, tools, tool_categories, matrix,
        health, extra_columns, agreement_cats, agreements,
        total_seconds, consolidated_csv, consolidated_json,
        cats_in_run=cats_in_run, cat_aggregates=cat_aggregates, holistic=holistic,
        structural=structural,
    )

    report_html = out_root / "REPORT.html"
    _render_report_html(
        report_html, fasta_in, peptides, tools, tool_categories, matrix,
        health, extra_columns, agreement_cats, agreements,
        total_seconds, consolidated_csv, consolidated_json, report_md,
        cats_in_run=cats_in_run, cat_aggregates=cat_aggregates, holistic=holistic,
        structural=structural,
    )

    consolidated_xlsx = out_root / "consolidated.xlsx"
    if _OPENPYXL_OK:
        _write_consolidated_xlsx(
            consolidated_xlsx, fasta_in, peptides, tools, tool_categories,
            matrix, extra_columns, agreement_cats, agreements, health,
            total_seconds,
            cats_in_run=cats_in_run, cat_aggregates=cat_aggregates, holistic=holistic,
            structural=structural,
        )

    print(f"[OK] consolidated.csv → {consolidated_csv}")
    print(f"[OK] consolidated.json → {consolidated_json}")
    if _OPENPYXL_OK:
        print(f"[OK] consolidated.xlsx → {consolidated_xlsx}")
    else:
        print("[WARN] openpyxl not installed — skipping consolidated.xlsx")
    print(f"[OK] tool_health_report.json → {health_report}")
    print(f"[OK] REPORT.md → {report_md}")
    print(f"[OK] REPORT.html → {report_html}")
    print(f"[OK] per-tool dir → {per_tool_root}")
    print(f"[OK] total: {total_seconds}s, {sum(1 for h in health.values() if h['status']=='OK')}/{len(tools)} OK")
    print(f"[OK] output dir → {out_root}")


if __name__ == "__main__":
    main()
